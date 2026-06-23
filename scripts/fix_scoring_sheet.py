"""
修正 评分表.xlsx：照搬 评分表20250601模板.xlsx 的填法
组长必填：组长/组员/贡献度/测试文档分数(40)/用例执行完成否/用例数量各项/最终得分
老师填：测试计划/说明/报告质量(14/13/13)、用例质量等级、用例质量(15)、测试类型数量(20)、问题回答等级、问题回答(10) ← 这些留空
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

p = r"D:/code/StarPicture/docs/test/评分表.xlsx"
wb = load_workbook(p)
ws = wb.active
ws.title = "评分表"

# 清空旧的黄/紫色填充（让它恢复成模板原本的样子）
white = PatternFill(fill_type=None)

# 数据行从第 3 行开始，共 4 行
for r in range(3, 7):
    for c_idx in range(1, 25):
        cell = ws.cell(row=r, column=c_idx)
        cell.fill = white

# 现在按原模板的样子重新涂色
# 黄色 = 组长必填列：1(组长) 2(组员) 3(贡献度) 9(用例执行完成否) 17(用例数量合计) 18(用例数量得分15) 23(备注) 24(最终得分)
# 紫色 = 老师填列：4(测试计划) 5(测试说明) 6(测试报告) 8(测试文档分数) 19(用例质量等级) 20(用例质量) 21(测试类型数量) 22(问题回答等级)
# 其余（7空列, 10-16各项用例类型数量）保持不涂色 ← 原模板也是白的
yellow_fill = PatternFill("solid", fgColor="FFF2CC")
purple_fill = PatternFill("solid", fgColor="E4D5F0")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 重置数据
data = [
    # 组长 朱远亮
    ["朱远亮", "朱远亮", 1.1, None, None, None, None, 16, "是",
     53, 1, 2, 3, 1, 2, 3, 65, 15, None, None, None, None, 4, "[=40+15+13+20+10]"],
    # 李冠燃
    [None, "李冠燃", 1.0, None, None, None, None, 16, "是",
     58, 2, 2, 3, 2, 2, 3, 72, 15, None, None, None, None, 4, "[=40+15+13+20+10]"],
    # 李坤纬
    [None, "李坤纬", 1.0, None, None, None, None, 16, "是",
     32, 1, 1, 2, 1, 1, 1, 39, 15, None, None, None, None, 4, "[=40+15+13+20+10]"],
    # 林景彬
    [None, "林景彬", 1.0, None, None, None, None, 16, "是",
     19, 1, 2, 3, 2, 2, 2, 31, 15, None, None, None, None, 4, "[=40+15+13+20+10]"],
]
# 写入
for r_idx, row in enumerate(data, start=3):
    for c_idx, v in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=v)
        cell.alignment = center
        cell.border = border

# 涂色
yellow_cols = {1, 2, 3, 9, 17, 18, 23, 24}
purple_cols = {4, 5, 6, 8, 19, 20, 21, 22}
for r in range(3, 7):
    for c_idx in range(1, 25):
        if c_idx in yellow_cols:
            ws.cell(row=r, column=c_idx).fill = yellow_fill
        elif c_idx in purple_cols:
            ws.cell(row=r, column=c_idx).fill = purple_fill

# 说明行（也照原模板）
for note_row in (7, 8):
    for c_idx in range(1, 25):
        cell = ws.cell(row=note_row, column=c_idx)
        cell.fill = yellow_fill
        cell.font = Font(italic=True, color="666666")
        cell.alignment = Alignment(horizontal="left", vertical="center")

ws.cell(row=7, column=1, value="黄色区域为组长填写区域，其余部分由老师填写")
ws.cell(row=8, column=1, value="淡紫色区域空着，由老师填写")

wb.save(p)
print(f"已修正: {p}")
print("黄色（组长填）: 1 组长/2 组员/3 贡献度/9 用例执行完成否/17 用例数量/18 用例数量得分/23 问题回答/24 最终得分")
print("紫色（老师填）: 4 测试计划/5 测试说明/6 测试报告/8 测试文档分数/19 用例质量等级/20 用例质量/21 测试类型数量/22 问题回答等级")
