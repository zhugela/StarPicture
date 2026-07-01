"""
重做！按老师要求：4 人 × 2 个功能点 + 4 种测试类型（功能/性能/接口/安全）
保留 4 种测试类型 = 测试类型数量 20 分能拿满分
保留 5 个子目录（性能/接口/安全/自动化/单元），但每个目录里只放与 2 个功能相关的最小集
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

BASE = Path("D:/code/StarPicture/docs/test")
PRODUCT = "内娱图库StarPicture"

# 通用样式
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="305496")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(border_style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["用例编号", "所属产品", "所属模块", "用例类型", "优先级", "用例标题",
           "前置条件", "步骤", "预期结果", "实测结果", "结论", "测试人员", "测试时间"]

# ============================================================
# 4 人 × 2 功能点 = 8 个功能点 + 每人 1 性能 + 1 接口 + 1 安全
# ============================================================
# 朱远亮：用户注册 + 用户登录
# 李冠燃：图片本地上传 + 关键字搜图
# 李坤纬：空间创建 + 空间成员管理
# 林景彬：文件上传 + 微信公众号门户

CASES = {
    "朱远亮": [
        # ====== 功能：用户注册 ======
        ("TC-UR-001", "/用户注册", "功能测试", 1, "注册_账号密码正确", "无账号",
         "1. POST /user/register userAccount=Zyl_New01 userPassword=12345678 checkPassword=12345678",
         "返回code=0, userId>0"),
        ("TC-UR-002", "/用户注册", "功能测试", 1, "注册_账号已存在", "TC_New01 已存在",
         "1. 重复 TC-UR-001",
         "返回code=40001 提示账号重复"),
        ("TC-UR-003", "/用户注册", "功能测试", 2, "注册_密码不一致", "无账号",
         "1. userPassword=12345678 checkPassword=87654321",
         "返回code=40001"),
        ("TC-UR-004", "/用户注册", "功能测试", 2, "注册_账号为空", "无账号",
         "1. userAccount=''",
         "返回code=40001"),
        ("TC-UR-005", "/用户注册", "功能测试", 2, "注册_账号长度不足4", "无账号",
         "1. userAccount='abc'(3字符)",
         "返回code=40001 提示账号长度不足"),
        # ====== 功能：用户登录 ======
        ("TC-UL-001", "/用户登录", "功能测试", 1, "登录_账号密码正确", "testuser01 已存在",
         "1. POST /user/login userAccount=testuser01 userPassword=12345678",
         "返回code=0, Set-Cookie 含登录态"),
        ("TC-UL-002", "/用户登录", "功能测试", 1, "登录_密码错误", "testuser01 已存在",
         "1. password=wrongpass",
         "返回code=40001"),
        ("TC-UL-003", "/用户登录", "功能测试", 2, "登录_账号不存在", "无",
         "1. userAccount=testuser99",
         "返回code=40001"),
        ("TC-UL-004", "/用户登录", "功能测试", 2, "登录_账号密码都为空", "无",
         "1. 空 body",
         "返回code=40001"),
        ("TC-UL-005", "/用户登录", "功能测试", 2, "登录_密码小于8字符", "无",
         "1. password='1234567'",
         "返回code=40001"),
        # ====== 性能：登录 50 并发 ======
        ("TC-PERF-001", "/用户登录", "性能测试", 2, "登录_50并发", "已安装 JMeter",
         "1. JMeter 50 线程 5s 内启动\n2. POST /user/login 正确账号密码",
         "P95 < 500ms, 错误率 < 1%"),
        # ====== 接口：Content-Type 错误 ======
        ("TC-API-001", "/用户注册", "接口测试", 2, "注册_Content-Type错误", "无账号",
         "1. Content-Type=application/xml\n2. POST /user/register body=XML",
         "返回 415"),
        # ====== 安全：SQL 注入 ======
        ("TC-SEC-001", "/用户登录", "安全测试", 1, "登录_SQL注入", "testuser01 已存在",
         "1. password = \"' OR 1=1 --\"\n2. POST /user/login",
         "返回code=40001 拒绝登录"),
        # ====== 扩展：注册边界/异常场景 =====
        ("TC-UR-006", "/用户注册", "功能测试", 2, "注册_账号长度边界20", "无账号",
         "1. userAccount=20字符",
         "返回code=0"),
        ("TC-UR-007", "/用户注册", "功能测试", 2, "注册_账号含特殊字符", "无账号",
         "1. userAccount='test@user.cn'",
         "返回code=40001"),
        ("TC-UR-008", "/用户注册", "功能测试", 2, "注册_纯数字账号", "无账号",
         "1. userAccount='123456'",
         "返回code=40001 纯数字不允许"),
        ("TC-UR-009", "/用户注册", "功能测试", 2, "注册_密码含中文", "无账号",
         "1. userPassword='密码12345678'",
         "返回code=40001"),
        ("TC-UR-010", "/用户注册", "功能测试", 2, "注册_账号含emoji", "无账号",
         "1. userAccount='test😀01'",
         "返回code=40001"),
        # ====== 扩展：登录异常/边界 =====
        ("TC-UL-006", "/用户登录", "功能测试", 1, "登录_退出后Cookie失效", "testuser01 已登录",
         "1. 先登录拿Cookie\n2. POST /user/logout\n3. 用旧Cookie访问 /user/get/login",
         "返回code=40100"),
        ("TC-UL-007", "/用户登录", "功能测试", 2, "登录_连续5次密码错", "testuser01 已存在",
         "1. 连续5次错误密码",
         "返回5次code=40001"),
        ("TC-UL-008", "/用户登录", "接口测试", 2, "登录_缺Content-Type", "无",
         "1. 不设置Content-Type\n2. POST /user/login",
         "返回415或服务端解析"),
        ("TC-UL-009", "/用户登录", "功能测试", 2, "登录_密码仅大小写不同", "testuser01 已存在",
         "1. password='12345678A' (实际是小写)",
         "返回code=40001"),
        ("TC-UL-010", "/用户登录", "功能测试", 2, "登录_密码含空格", "testuser01 已存在",
         "1. password=' 12345678 '",
         "返回code=40001"),
        # ====== 扩展：管理员操作 =====
        ("TC-UA-001", "/用户管理", "功能测试", 2, "管理员查询用户列表", "admin 已登录",
         "1. POST /user/list/page/vo current=1 pageSize=10",
         "返回code=0, records含用户列表"),
        ("TC-UA-002", "/用户管理", "功能测试", 2, "管理员按昵称搜索", "admin 已登录",
         "1. POST /user/list/page/vo userName='test'",
         "返回code=0 含test的记录"),
        ("TC-UA-003", "/用户管理", "安全测试", 1, "普通用户越权调用", "testuser01 已登录",
         "1. POST /user/list/page/vo",
         "返回code=40300"),
        # ====== 扩展：性能 =====
        ("TC-PERF-002", "/用户注册", "性能测试", 2, "注册_20并发", "已安装 JMeter",
         "1. JMeter 20 线程 5s 内启动\n2. POST /user/register",
         "P95 < 800ms, 错误率 < 1%"),
        # ====== 扩展：接口契约 =====
        ("TC-API-002", "/用户注册", "接口测试", 2, "注册_缺userAccount字段", "无账号",
         "1. body 只有 userPassword 和 checkPassword",
         "返回code=40001"),
    ],

    "李冠燃": [
        # ====== 功能：本地上传 ======
        ("TC-PU-001", "/图片上传", "功能测试", 1, "本地上传_jpg_2MB", "testuser01 已登录",
         "1. POST /file/upload file=2MB_jpg",
         "返回code=0, url 字段为可访问图片链接"),
        ("TC-PU-002", "/图片上传", "功能测试", 1, "本地上传_png", "testuser01 已登录",
         "1. POST /file/upload file=1MB_png",
         "返回code=0, picFormat=png"),
        ("TC-PU-003", "/图片上传", "功能测试", 1, "本地上传_超过2MB_应失败", "testuser01 已登录",
         "1. POST /file/upload file=5MB_jpg",
         "返回code=40001"),
        ("TC-PU-004", "/图片上传", "功能测试", 2, "本地上传_非图片_应失败", "testuser01 已登录",
         "1. POST /file/upload file=test.pdf",
         "返回code=40001"),
        ("TC-PU-005", "/图片上传", "功能测试", 2, "本地上传_空文件_应失败", "testuser01 已登录",
         "1. POST /file/upload file=0byte",
         "返回code=40001"),
        # ====== 功能：关键字搜索 ======
        ("TC-PX-001", "/图片搜索", "功能测试", 1, "关键字搜图_有结果", "已有图片",
         "1. POST /picture/search/picture text='猫'",
         "返回含'猫'的图片列表"),
        ("TC-PX-002", "/图片搜索", "功能测试", 1, "关键字搜图_无结果", "无",
         "1. POST /picture/search/picture text='xyz999'",
         "返回空列表"),
        ("TC-PX-003", "/图片搜索", "功能测试", 2, "关键字搜图_空文本", "testuser01 已登录",
         "1. POST /picture/search/picture text=''",
         "返回code=40001 或全部图片"),
        ("TC-PX-004", "/图片搜索", "功能测试", 2, "关键字搜图_超长", "testuser01 已登录",
         "1. text=1000字符",
         "返回code=40001"),
        ("TC-PX-005", "/图片搜索", "功能测试", 2, "关键字搜图_未登录", "无",
         "1. POST /picture/search/picture",
         "返回code=40100"),
        # ====== 性能：上传 20 并发 ======
        ("TC-PERF-001", "/图片上传", "性能测试", 2, "上传_20并发", "已安装 JMeter",
         "1. JMeter 20 线程上传 2MB jpg",
         "P95 < 3s, 错误率 < 5%"),
        # ====== 接口：multipart 边界 ======
        ("TC-API-001", "/图片上传", "接口测试", 2, "上传_缺multipart边界", "testuser01 已登录",
         "1. 发送不完整 multipart 请求",
         "返回 400 或 415"),
        # ====== 安全：木马上传 ======
        ("TC-SEC-001", "/图片上传", "安全测试", 1, "上传_伪PHP木马", "testuser01 已登录",
         "1. file=1.jpg 内容为 <?php system($_GET['c']);?>",
         "返回code=40001"),
        # ====== 扩展：上传格式/边界 =====
        ("TC-PU-006", "/图片上传", "功能测试", 2, "上传_bmp格式", "testuser01 已登录",
         "1. POST /file/upload file=1MB_bmp",
         "返回code=0 picFormat=bmp"),
        ("TC-PU-007", "/图片上传", "功能测试", 2, "上传_gif格式", "testuser01 已登录",
         "1. POST /file/upload file=1MB_gif",
         "返回code=0 picFormat=gif"),
        ("TC-PU-008", "/图片上传", "功能测试", 2, "上传_webp格式", "testuser01 已登录",
         "1. POST /file/upload file=1MB_webp",
         "返回code=0 picFormat=webp"),
        ("TC-PU-009", "/图片上传", "功能测试", 2, "上传_文件名含中文", "testuser01 已登录",
         "1. POST /file/upload file=中文图片.jpg",
         "返回code=0 url 正确编码"),
        ("TC-PU-010", "/图片上传", "功能测试", 2, "上传_文件名含特殊字符", "testuser01 已登录",
         "1. POST /file/upload file='im$g.jpg'",
         "返回code=0 url 已转义"),
        # ====== 扩展：搜图异常 =====
        ("TC-PX-006", "/图片搜索", "功能测试", 1, "搜图_精确匹配", "已有图片",
         "1. POST /picture/search/picture text='StarPicture'",
         "返回code=0 含结果"),
        ("TC-PX-007", "/图片搜索", "接口测试", 2, "搜图_缺text字段", "testuser01 已登录",
         "1. body 为空",
         "返回code=40001"),
        ("TC-PX-008", "/图片搜索", "安全测试", 1, "搜图_SQL注入", "testuser01 已登录",
         "1. text=\"' OR 1=1 --\"",
         "返回code=40001 拒绝注入"),
        ("TC-PX-009", "/图片搜索", "安全测试", 1, "搜图_XSS注入", "testuser01 已登录",
         "1. text=\"<script>alert(1)</script>\"",
         "返回code=0 渲染时被转义"),
        ("TC-PX-010", "/图片搜索", "功能测试", 2, "搜图_含Unicode字符", "已有图片",
         "1. text='测试中文搜索'",
         "返回code=0"),
        # ====== 扩展：图片管理 =====
        ("TC-PG-001", "/图片管理", "功能测试", 1, "按id查询图片", "已有图片",
         "1. GET /picture/get/vo?id=1",
         "返回code=0 含完整图片信息"),
        ("TC-PG-002", "/图片管理", "功能测试", 1, "分页查询所有图片", "testuser01 已登录",
         "1. POST /picture/list/page/vo current=1 pageSize=10",
         "返回code=0 records含图片列表"),
        ("TC-PG-003", "/图片管理", "接口测试", 2, "查询_id类型错误", "testuser01 已登录",
         "1. GET /picture/get/vo?id='abc'",
         "返回code=40001"),
        # ====== 扩展：性能 =====
        ("TC-PERF-002", "/图片搜索", "性能测试", 2, "搜图_50并发", "已安装 JMeter",
         "1. JMeter 50 线程 POST /picture/search/picture",
         "P95 < 500ms, 错误率 < 1%"),
        # ====== 扩展：URL 上传扩展 =====
        ("TC-URL-001", "/图片上传", "安全测试", 1, "URL上传_内网IP", "testuser01 已登录",
         "1. fileUrl='http://127.0.0.1/x.jpg'",
         "返回code=40001 拒绝内网"),
    ],

    "李坤纬": [
        # ====== 功能：创建空间 ======
        ("TC-SP-001", "/空间管理", "功能测试", 1, "获取空间等级列表", "无",
         "1. GET /space/list/level",
         "返回code=0, 含普通版/专业版/旗舰版"),
        ("TC-SP-002", "/空间管理", "功能测试", 1, "创建空间_普通版", "testuser01 已登录",
         "1. POST /space/add spaceName='我的空间' spaceLevel=0",
         "返回code=0, spaceId 存在"),
        ("TC-SP-003", "/空间管理", "功能测试", 2, "创建空间_名称为空", "testuser01 已登录",
         "1. POST /space/add spaceName=''",
         "返回code=40001"),
        ("TC-SP-004", "/空间管理", "功能测试", 2, "创建空间_名称超长", "testuser01 已登录",
         "1. POST /space/add spaceName=50+'x'",
         "返回code=40001"),
        ("TC-SP-005", "/空间管理", "功能测试", 2, "创建空间_未登录", "无",
         "1. POST /space/add",
         "返回code=40100"),
        # ====== 功能：空间成员管理 ======
        ("TC-SU-001", "/空间成员", "功能测试", 1, "添加空间成员", "admin 已登录, testuser01 有空间",
         "1. POST /spaceUser/add spaceId=1 userId=2 spaceRole='viewer'",
         "返回code=0"),
        ("TC-SU-002", "/空间成员", "功能测试", 1, "查询空间成员列表", "admin 已登录",
         "1. POST /spaceUser/list spaceId=1",
         "返回成员列表"),
        ("TC-SU-003", "/空间成员", "功能测试", 2, "添加成员_重复", "成员已存在",
         "1. 重复 TC-SU-001",
         "返回code=40001"),
        ("TC-SU-004", "/空间成员", "功能测试", 2, "添加成员_未登录", "无",
         "1. POST /spaceUser/add",
         "返回code=40100"),
        ("TC-SU-005", "/空间成员", "功能测试", 2, "删除空间成员", "admin 已登录",
         "1. POST /spaceUser/delete id=1",
         "返回code=0, isDelete=1"),
        # ====== 性能：空间分析 20 并发 ======
        ("TC-PERF-001", "/空间分析", "性能测试", 2, "空间分析_20并发", "已安装 JMeter",
         "1. JMeter 20 线程 POST /space/analyze/usage",
         "P95 < 1s, 错误率 < 1%"),
        # ====== 接口：缺 Content-Type ======
        ("TC-API-001", "/空间管理", "接口测试", 2, "创建空间_缺Content-Type", "testuser01 已登录",
         "1. 不设置 Content-Type\n2. POST /space/add JSON body",
         "返回 415 或服务端能解析"),
        # ====== 安全：空间名称 XSS ======
        ("TC-SEC-001", "/空间管理", "安全测试", 1, "空间名称XSS", "testuser01 已登录",
         "1. POST /space/add spaceName='<script>alert(1)</script>'",
         "返回code=0, 渲染时应被转义"),
        # ====== 扩展：创建边界 =====
        ("TC-SP-006", "/空间管理", "功能测试", 2, "创建_重复同名空间", "testuser01 已登录",
         "1. 重复创建相同 spaceName",
         "返回code=40001 或 code=0(允许同名)"),
        ("TC-SP-007", "/空间管理", "功能测试", 2, "创建_名称50字符", "testuser01 已登录",
         "1. spaceName=50+'x'",
         "返回code=40001 超长"),
        ("TC-SP-008", "/空间管理", "功能测试", 2, "创建_名称含特殊字符", "testuser01 已登录",
         "1. spaceName='my space!@#'",
         "返回code=40001"),
        ("TC-SP-009", "/空间管理", "功能测试", 2, "创建_名称含中文", "testuser01 已登录",
         "1. spaceName='我的空间'",
         "返回code=0 支持中文"),
        ("TC-SP-010", "/空间管理", "功能测试", 2, "创建_旗舰版spaceLevel=3", "testuser01 已登录",
         "1. spaceName='旗舰' spaceLevel=3",
         "返回code=0"),
        # ====== 扩展：成员扩展 =====
        ("TC-SU-006", "/空间成员", "功能测试", 2, "成员_重复添加", "admin 已登录",
         "1. 重复添加同一成员",
         "返回code=40001"),
        ("TC-SU-007", "/空间成员", "功能测试", 2, "成员_添加自己", "testuser01 已登录",
         "1. userId=当前用户自己",
         "返回code=40001 不能加自己"),
        ("TC-SU-008", "/空间成员", "功能测试", 2, "成员_查询不存在的空间", "admin 已登录",
         "1. POST /spaceUser/list spaceId=999999",
         "返回空列表"),
        ("TC-SU-009", "/空间成员", "接口测试", 2, "成员_缺spaceId", "admin 已登录",
         "1. body 无 spaceId",
         "返回code=40001"),
        ("TC-SU-010", "/空间成员", "功能测试", 2, "成员_删除已删除的成员", "admin 已登录",
         "1. 重复 delete 同一成员",
         "返回code=40001 第二次失败"),
        # ====== 扩展：空间分析扩展 =====
        ("TC-SA-001", "/空间分析", "功能测试", 1, "空间分类分析", "testuser01 有空间",
         "1. POST /space/analyze/category spaceId=1",
         "返回code=0 分类统计"),
        ("TC-SA-002", "/空间分析", "功能测试", 1, "空间标签分析", "testuser01 有空间",
         "1. POST /space/analyze/tag spaceId=1",
         "返回code=0 标签统计"),
        ("TC-SA-003", "/空间分析", "功能测试", 1, "空间排行", "testuser01 已登录",
         "1. POST /space/analyze/rank",
         "返回code=0 排行列表"),
        # ====== 扩展：安全 =====
        ("TC-SEC-002", "/空间管理", "安全测试", 1, "跨用户删除空间", "testuser02 已登录",
         "1. POST /space/delete id=testuser01的空间",
         "返回code=40300 越权失败"),
        # ====== 扩展：性能 =====
        ("TC-PERF-002", "/空间列表", "性能测试", 2, "空间列表_50并发", "已安装 JMeter",
         "1. JMeter 50 线程 POST /space/list/page/vo",
         "P95 < 500ms, 错误率 < 1%"),
        # ====== 扩展：接口契约 =====
        ("TC-API-002", "/空间成员", "接口测试", 2, "添加成员_缺userId", "admin 已登录",
         "1. body 只有 spaceId 和 spaceRole",
         "返回code=40001"),
    ],

    "林景彬": [
        # ====== 功能：本地上传 ======
        ("TC-FL-001", "/文件上传", "功能测试", 1, "本地上传_jpg_2MB", "testuser01 已登录",
         "1. POST /file/upload file=2MB_jpg",
         "返回code=0, url 字段为可访问图片链接"),
        ("TC-FL-002", "/文件上传", "功能测试", 1, "本地上传_png", "testuser01 已登录",
         "1. POST /file/upload file=1MB_png",
         "返回code=0"),
        ("TC-FL-003", "/文件上传", "功能测试", 1, "本地上传_avatar_头像", "testuser01 已登录",
         "1. POST /file/upload/avatar file=512KB_jpg",
         "返回code=0, userAvatar 已更新"),
        ("TC-FL-004", "/文件上传", "功能测试", 2, "本地上传_超过2MB_应失败", "testuser01 已登录",
         "1. POST /file/upload file=5MB_jpg",
         "返回code=40001"),
        ("TC-FL-005", "/文件上传", "功能测试", 2, "本地上传_未登录", "无",
         "1. POST /file/upload",
         "返回code=40100"),
        # ====== 功能：微信公众号门户 ======
        ("TC-WX-001", "/微信公众号", "功能测试", 1, "门户_GET_签名验证", "微信服务器 GET",
         "1. GET /wx/mp/portal signature=xxx timestamp=xxx nonce=xxx echostr=xxx",
         "返回 echostr 明文（签名验证通过）"),
        ("TC-WX-002", "/微信公众号", "功能测试", 1, "门户_POST_消息处理", "微信服务器 POST XML",
         "1. POST /wx/mp/portal Content-Type=application/xml\n2. body=<xml>...</xml>",
         "返回code=0, 响应 XML 包含正确结构"),
        ("TC-WX-003", "/微信公众号", "功能测试", 2, "创建公众号菜单", "admin 已登录",
         "1. POST /wx/mp/menu/create menu={\"button\":[...]}",
         "返回code=0"),
        ("TC-WX-004", "/微信公众号", "功能测试", 2, "门户_GET_签名错误", "伪造签名",
         "1. GET /wx/mp/portal signature=invalid",
         "返回空或失败提示"),
        ("TC-WX-005", "/微信公众号", "功能测试", 2, "门户_POST_未登录", "无",
         "1. POST /wx/mp/portal",
         "返回code=40100 或 40300"),
        # ====== 性能：上传 50 并发 ======
        ("TC-PERF-001", "/文件上传", "性能测试", 2, "上传_50并发_1MB", "已安装 JMeter",
         "1. JMeter 50 线程上传 1MB jpg",
         "P95 < 2s, 错误率 < 5%"),
        # ====== 接口：缺 multipart 边界 ======
        ("TC-API-001", "/文件上传", "接口测试", 2, "upload_缺multipart边界", "testuser01 已登录",
         "1. 发送不完整 multipart 请求",
         "返回 400 或 415"),
        # ====== 安全：双扩展名绕过 ======
        ("TC-SEC-001", "/文件上传", "安全测试", 1, "上传_双扩展名", "testuser01 已登录",
         "1. file=1.jpg.php",
         "返回code=40001"),
        # ====== 扩展：头像上传格式 =====
        ("TC-AT-006", "/头像上传", "功能测试", 2, "头像_超过1MB_应失败", "testuser01 已登录",
         "1. POST /file/upload/avatar file=1.5MB_jpg",
         "返回code=40001 超限"),
        ("TC-AT-007", "/头像上传", "功能测试", 2, "头像_非图片文件_应失败", "testuser01 已登录",
         "1. POST /file/upload/avatar file=test.pdf",
         "返回code=40001"),
        ("TC-AT-008", "/头像上传", "功能测试", 2, "头像_webp格式", "testuser01 已登录",
         "1. POST /file/upload/avatar file=300KB_webp",
         "返回code=0 picFormat=webp"),
        ("TC-AT-009", "/头像上传", "功能测试", 2, "头像_空文件_应失败", "testuser01 已登录",
         "1. POST /file/upload/avatar file=0byte",
         "返回code=40001"),
        ("TC-AT-010", "/头像上传", "功能测试", 2, "头像_重复上传覆盖旧头像", "testuser01 已登录",
         "1. 先上传头像A\n2. 再上传头像B",
         "返回code=0 userAvatar 改为新URL"),
        # ====== 扩展：获取当前用户扩展 =====
        ("TC-GC-006", "/获取当前用户", "功能测试", 1, "获取_管理员视角", "admin 已登录",
         "1. GET /user/get/login (admin Cookie)",
         "返回code=0 含 userRole=admin"),
        ("TC-GC-007", "/获取当前用户", "接口测试", 2, "获取_id为字符串", "无",
         "1. Cookie伪造 userId='abc'",
         "返回code=40100"),
        ("TC-GC-008", "/获取当前用户", "功能测试", 1, "获取_用户信息完整", "testuser01 已登录",
         "1. GET /user/get/login",
         "返回code=0 含 userAvatar/userName/userProfile"),
        ("TC-GC-009", "/获取当前用户", "安全测试", 1, "XSS_用户简介", "testuser01 已登录",
         "1. POST /user/update/my userProfile='<script>alert(1)</script>'",
         "返回code=0 渲染时被转义"),
        ("TC-GC-010", "/获取当前用户", "功能测试", 1, "获取_用户简介含emoji", "testuser01 已登录",
         "1. POST /user/update/my userProfile='😀 用户'",
         "返回code=0"),
        # ====== 扩展：用户更新 =====
        ("TC-UU-001", "/用户更新", "功能测试", 1, "更新_昵称", "testuser01 已登录",
         "1. POST /user/update/my userName='新名字'",
         "返回code=0"),
        ("TC-UU-002", "/用户更新", "功能测试", 2, "更新_昵称为空", "testuser01 已登录",
         "1. POST /user/update/my userName=''",
         "返回code=40001"),
        ("TC-UU-003", "/用户更新", "接口测试", 2, "更新_缺userName", "testuser01 已登录",
         "1. body 无 userName",
         "返回code=40001"),
        # ====== 扩展：性能 =====
        ("TC-PERF-002", "/头像上传", "性能测试", 2, "头像上传_100并发", "已安装 JMeter",
         "1. JMeter 100 线程 POST /file/upload/avatar",
         "P95 < 3s, 错误率 < 5%"),
        # ====== 扩展：头像安全 =====
        ("TC-SEC-002", "/头像上传", "安全测试", 1, "头像_伪SVG含JS", "testuser01 已登录",
         "1. file=avatar.svg 内容含 <script>alert(1)</script>",
         "返回code=40001 拒绝SVG"),
    ],
}

# 统计并打印
print("=== 重新生成 4 人用例（每人 2 个功能 + 1 性能 + 1 接口 + 1 安全）===")
total_all = 0
total_by_type = {"功能测试": 0, "性能测试": 0, "接口测试": 0, "安全测试": 0}
for m, cases in CASES.items():
    sub = len(cases)
    total_all += sub
    t_count = {}
    for c in cases:
        t_count[c[2]] = t_count.get(c[2], 0) + 1
    print(f"\n{m}: {sub} 条")
    for t, n in sorted(t_count.items()):
        total_by_type[t] = total_by_type.get(t, 0) + n
        print(f"  {t}: {n}")
print(f"\n总计: {total_all} 条")
print("按类型:", total_by_type)

# ============================================================
# 生成 4 份 xlsx
# ============================================================
def build_xlsx(member, cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append(HEADERS)
    for c_idx in range(1, len(HEADERS)+1):
        cell = ws.cell(row=1, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for i, c in enumerate(cases, start=1):
        row = [c[0], PRODUCT, c[1], c[2], c[3], c[4], c[5], c[6], c[7], "", "", member, ""]
        ws.append(row)

    for r in range(2, ws.max_row+1):
        for c_idx in range(1, len(HEADERS)+1):
            cell = ws.cell(row=r, column=c_idx)
            cell.alignment = left
            cell.border = border
            if c_idx in (3, 4, 5):
                cell.alignment = center
            if c_idx == 6:
                cell.font = Font(bold=True)

    col_widths = [14, 18, 16, 14, 6, 30, 28, 38, 32, 14, 8, 10, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    for r in range(2, ws.max_row+1):
        ws.row_dimensions[r].height = 80
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{ws.max_row}"

    # 统计 sheet
    ws2 = wb.create_sheet("用例统计")
    ws2.append(["测试类型", "用例数量"])
    t_count = {}
    for c in cases:
        t_count[c[2]] = t_count.get(c[2], 0) + 1
    for k, v in sorted(t_count.items(), key=lambda x: -x[1]):
        ws2.append([k, v])
    ws2.append(["合计", len(cases)])
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 12
    for r in range(1, ws2.max_row+1):
        for c_idx in range(1, 3):
            cell = ws2.cell(row=r, column=c_idx)
            cell.alignment = center
            cell.border = border
            if r == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill

    out_path = f"{BASE}/{member}_脚本与截图/软件测试测试用例.xlsx"
    wb.save(out_path)
    return out_path, len(cases), t_count

for m, cases in CASES.items():
    path, n, tc = build_xlsx(m, cases)
    print(f"\n[{m}] -> {n} 条 -> {path}")
