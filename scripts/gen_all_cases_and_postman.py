"""
一次性生成：
1. 4份分模块 xlsx（每份的用例完全对应实际 Controller API）
2. 1份汇总 xlsx
3. 4个 Postman JSON（可导入运行）
4. 更新评分表
"""
import json
import shutil
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = Path(r"D:\code\StarPicture\docs\内娱图库_海蒂与爷爷_朱远亮_18144610287")
SCRIPTS = Path(r"D:\code\StarPicture\scripts")
PRODUCT = "StarPicture"

# ============================================================
# 1. 每个成员的真实 API 测试用例（完全按 Controller 来）
# ============================================================

CASES = {
    "朱远亮": [
        # ---- /user/register ----
        ("TC-UR-001", "/user/register", "功能测试", 1, "注册_正常", "无",
         "POST /user/register\nBody: userAccount=newtest, userPassword=Test12345, checkPassword=Test12345",
         "code=0, data 含 userId"),
        ("TC-UR-002", "/user/register", "功能测试", 1, "注册_账号已存在", "newtest 已注册",
         "POST /user/register\nBody: userAccount=newtest, userPassword=Test12345, checkPassword=Test12345",
         "code=40001, 提示账号已存在"),
        ("TC-UR-003", "/user/register", "功能测试", 2, "注册_账号为空", "无",
         "POST /user/register\nBody: userAccount=, userPassword=Test12345, checkPassword=Test12345",
         "code=40001, 提示账号不能为空"),
        ("TC-UR-004", "/user/register", "功能测试", 2, "注册_密码不一致", "无",
         "POST /user/register\nBody: userAccount=test01, userPassword=Test12345, checkPassword=Test11111",
         "code=40001, 提示密码不一致"),
        ("TC-UR-005", "/user/register", "功能测试", 2, "注册_账号长度不足4", "无",
         "POST /user/register\nBody: userAccount=abc, userPassword=Test12345, checkPassword=Test12345",
         "code=40001, 提示账号长度不足"),
        # ---- /user/login ----
        ("TC-UL-001", "/user/login", "功能测试", 1, "登录_正常", "testuser01 已存在",
         "POST /user/login\nBody: userAccount=testuser01, userPassword=12345678",
         "code=0, Set-Cookie, data 含 userInfo"),
        ("TC-UL-002", "/user/login", "功能测试", 1, "登录_密码错误", "testuser01 已存在",
         "POST /user/login\nBody: userAccount=testuser01, userPassword=wrongpass",
         "code=40001, 提示密码错误"),
        ("TC-UL-003", "/user/login", "功能测试", 2, "登录_账号不存在", "无",
         "POST /user/login\nBody: userAccount=nonexist, userPassword=12345678",
         "code=40001, 提示账号不存在"),
        ("TC-UL-004", "/user/login", "安全测试", 1, "登录_SQL注入", "testuser01 已存在",
         "POST /user/login\nBody: userAccount=testuser01' OR '1'='1, userPassword=12345678",
         "code=40001, 拒绝登录"),
        ("TC-UL-005", "/user/login", "功能测试", 2, "登录_账号为空", "无",
         "POST /user/login\nBody: userAccount=, userPassword=12345678",
         "code=40001, 提示账号不能为空"),
        ("TC-UL-006", "/user/login", "功能测试", 2, "登录_密码为空", "无",
         "POST /user/login\nBody: userAccount=testuser01, userPassword=",
         "code=40001, 提示密码不能为空"),
        # ---- /user/get/login ----
        ("TC-GC-001", "/user/get/login", "功能测试", 1, "获取当前登录用户_已登录", "testuser01 Cookie",
         "GET /user/get/login\nCookie: xxx",
         "code=0, data 含 id, userAccount, userName"),
        ("TC-GC-002", "/user/get/login", "功能测试", 1, "获取当前登录用户_未登录", "无 Cookie",
         "GET /user/get/login",
         "code=40100, 未登录"),
        ("TC-GC-003", "/user/get/login", "安全测试", 1, "获取当前用户_伪造Cookie", "伪造的 Cookie",
         "GET /user/get/login\nCookie: jwt=xxx伪造的xxx",
         "code=40100, 未登录"),
        # ---- /user/update/my ----
        ("TC-UM-001", "/user/update/my", "功能测试", 1, "修改个人信息_正常", "testuser01 Cookie",
         "POST /user/update/my\nBody: userName=新昵称, userProfile=新简介",
         "code=0, data 含更新后的信息"),
        ("TC-UM-002", "/user/update/my", "功能测试", 2, "修改个人信息_未登录", "无 Cookie",
         "POST /user/update/my\nBody: userName=新昵称",
         "code=40100"),
        ("TC-UM-003", "/user/update/my", "安全测试", 1, "修改个人信息_XSS", "testuser01 Cookie",
         "POST /user/update/my\nBody: userName=<script>alert(1)</script>",
         "code=40001 或 userName 被过滤"),
        # ---- /user/add (admin) ----
        ("TC-UA-001", "/user/add", "功能测试", 1, "管理员新增用户", "admin Cookie",
         "POST /user/add\nBody: userAccount=newbyadmin, userPassword=Test12345",
         "code=0, data 含 userId"),
        ("TC-UA-002", "/user/add", "安全测试", 1, "新增用户_普通用户越权", "testuser01 Cookie",
         "POST /user/add\nBody: userAccount=xxx, userPassword=Test12345",
         "code=40300, 无权限"),
        # ---- /user/get ----
        ("TC-UG-001", "/user/get", "功能测试", 1, "按ID查询用户", "admin Cookie",
         "GET /user/get?id=1",
         "code=0, data 含 id=1 的用户信息"),
        ("TC-UG-002", "/user/get", "功能测试", 2, "按ID查询_用户不存在", "admin Cookie",
         "GET /user/get?id=99999",
         "code=40001, 用户不存在"),
        # ---- /user/get/vo ----
        ("TC-UGV-001", "/user/get/vo", "功能测试", 1, "按ID查询用户VO", "admin Cookie",
         "GET /user/get/vo?id=1",
         "code=0, data 含 UserVO"),
        # ---- /user/delete ----
        ("TC-UD-001", "/user/delete", "功能测试", 1, "删除用户_正常", "admin Cookie",
         "POST /user/delete\nBody: id=99",
         "code=0"),
        ("TC-UD-002", "/user/delete", "安全测试", 1, "删除用户_普通用户越权", "testuser01 Cookie",
         "POST /user/delete\nBody: id=2",
         "code=40300"),
        # ---- /user/update ----
        ("TC-UU-001", "/user/update", "功能测试", 1, "管理员更新用户", "admin Cookie",
         "POST /user/update\nBody: id=2, userName=updated",
         "code=0"),
        # ---- /user/list/page/vo ----
        ("TC-ULP-001", "/user/list/page/vo", "功能测试", 1, "分页查询用户", "admin Cookie",
         "POST /user/list/page/vo\nBody: current=1, pageSize=10",
         "code=0, records 有数据"),
        ("TC-ULP-002", "/user/list/page/vo", "功能测试", 2, "分页查询_第0页", "admin Cookie",
         "POST /user/list/page/vo\nBody: current=0, pageSize=10",
 "code=40001 或返回第1页"),
        # ---- /user/logout ----
        ("TC-ULO-001", "/user/logout", "功能测试", 1, "用户注销", "testuser01 Cookie",
         "POST /user/logout",
         "code=0, Cookie 被清除"),
        ("TC-ULO-002", "/user/logout", "功能测试", 2, "注销_未登录", "无 Cookie",
         "POST /user/logout",
         "code=40100"),
        # ---- 性能 ----
        ("TC-PERF-001", "/user/login", "性能测试", 2, "登录_50并发", "testuser01 已存在",
         "POST /user/login 50并发\nBody: userAccount=testuser01, userPassword=12345678",
         "P95 < 500ms, 错误率 < 5%"),
        # ---- 接口 ----
        ("TC-API-001", "/user/register", "接口测试", 2, "注册_Content-Type错误", "无",
         "POST /user/register\nHeader: Content-Type: application/xml\nBody: <xml>test</xml>",
         "code=40001 或 415"),
    ],

    "李冠燃": [
        # ---- /picture/upload (via /file/upload) ----
        ("TC-PU-001", "/picture/upload", "功能测试", 1, "本地上传_jpg_2MB", "testuser01 Cookie",
         "POST /file/upload\nBody: multipart file=2MB_jpg",
         "code=0, url 字段为可访问链接"),
        ("TC-PU-002", "/picture/upload", "功能测试", 1, "本地上传_png", "testuser01 Cookie",
         "POST /file/upload\nBody: multipart file=1MB_png",
         "code=0, picFormat=png"),
        ("TC-PU-003", "/picture/upload", "功能测试", 2, "本地上传_超过2MB", "testuser01 Cookie",
         "POST /file/upload\nBody: multipart file=5MB_jpg",
         "code=40001"),
        ("TC-PU-004", "/picture/upload", "功能测试", 2, "本地上传_非图片文件", "testuser01 Cookie",
         "POST /file/upload\nBody: multipart file=test.pdf",
         "code=40001"),
        ("TC-PU-005", "/picture/upload", "功能测试", 2, "本地上传_空文件", "testuser01 Cookie",
         "POST /file/upload\nBody: multipart file=0byte.jpg",
         "code=40001"),
        # ---- /picture/upload/url ----
        ("TC-PUU-001", "/picture/upload/url", "功能测试", 1, "URL上传_合法URL", "testuser01 Cookie",
         "POST /picture/upload/url\nBody: {\"fileUrl\":\"https://example.com/1.jpg\",\"picName\":\"test\"}",
         "code=0, picture.url 已写入"),
        ("TC-PUU-002", "/picture/upload/url", "安全测试", 1, "URL上传_内网URL_SSRF", "testuser01 Cookie",
         "POST /picture/upload/url\nBody: {\"fileUrl\":\"http://127.0.0.1/x.jpg\"}",
         "code=40001, 防止 SSRF"),
        ("TC-PUU-003", "/picture/upload/url", "功能测试", 2, "URL上传_404", "testuser01 Cookie",
         "POST /picture/upload/url\nBody: {\"fileUrl\":\"https://example.com/notfound.jpg\"}",
         "code=40001"),
        ("TC-PUU-004", "/picture/upload/url", "功能测试", 2, "URL上传_非图片URL", "testuser01 Cookie",
         "POST /picture/upload/url\nBody: {\"fileUrl\":\"https://example.com/index.html\"}",
         "code=40001"),
        # ---- /picture/upload/batch ----
        ("TC-PUB-001", "/picture/upload/batch", "功能测试", 1, "批量上传_5张", "testuser01 Cookie",
         "POST /picture/upload/batch\nBody: multipart files=[5张jpg]",
         "code=0, 5 条记录"),
        ("TC-PUB-002", "/picture/upload/batch", "功能测试", 2, "批量上传_超过20张", "testuser01 Cookie",
         "POST /picture/upload/batch\nBody: multipart files=[30张]",
         "code=40001 或只保存前N张"),
        # ---- /picture/get ----
        ("TC-PG-001", "/picture/get", "功能测试", 1, "按id查询原数据", "testuser01 Cookie",
         "GET /picture/get?id=1",
         "code=0, 含完整 url"),
        ("TC-PG-002", "/picture/get", "功能测试", 2, "查询_不存在id", "testuser01 Cookie",
         "GET /picture/get?id=999999",
         "code=40001"),
        # ---- /picture/get/vo ----
        ("TC-PGV-001", "/picture/get/vo", "功能测试", 1, "按id查询VO", "testuser01 Cookie",
         "GET /picture/get/vo?id=1",
         "code=0, PictureVO"),
        # ---- /picture/list/page ----
        ("TC-PLP-001", "/picture/list/page", "功能测试", 1, "分页查询_默认", "testuser01 Cookie",
         "POST /picture/list/page\nBody: {\"current\":1,\"pageSize\":10}",
         "code=0, 10 条记录"),
        ("TC-PLP-002", "/picture/list/page", "功能测试", 2, "分页_超大pageSize", "testuser01 Cookie",
         "POST /picture/list/page\nBody: {\"current\":1,\"pageSize\":10000}",
         "被限制或code=40001"),
        ("TC-PLP-003", "/picture/list/page", "功能测试", 2, "分页_按标签过滤", "testuser01 Cookie",
         "POST /picture/list/page\nBody: {\"tags\":[\"tag1\"]}",
         "返回含 tag1 的图片"),
        # ---- /picture/list/page/vo ----
        ("TC-PLPV-001", "/picture/list/page/vo", "功能测试", 1, "分页查询VO_管理员查所有", "admin Cookie",
         "POST /picture/list/page/vo\nBody: {\"current\":1,\"pageSize\":10}",
         "返回所有用户图片"),
        # ---- /picture/list/page/vo/cache ----
        ("TC-PLC-001", "/picture/list/page/vo/cache", "功能测试", 1, "缓存分页查询", "testuser01 Cookie",
         "POST /picture/list/page/vo/cache\nBody: {\"current\":1,\"pageSize\":10}",
         "code=0, 第二次更快"),
        # ---- /picture/delete ----
        ("TC-PD-001", "/picture/delete", "功能测试", 1, "删除自己的图片", "testuser01 Cookie",
         "POST /picture/delete\nBody: {\"id\":1}",
         "code=0, isDelete=1"),
        ("TC-PD-002", "/picture/delete", "安全测试", 1, "删除别人的图片_越权", "testuser02 Cookie",
         "POST /picture/delete\nBody: {\"id\":1}",
         "code=40300"),
        ("TC-PD-003", "/picture/delete", "功能测试", 1, "管理员删任意图片", "admin Cookie",
         "POST /picture/delete\nBody: {\"id\":2}",
         "code=0"),
        # ---- /picture/edit ----
        ("TC-PE-001", "/picture/edit", "功能测试", 1, "编辑名称简介标签", "testuser01 Cookie",
         "POST /picture/edit\nBody: {\"id\":1,\"name\":\"新名\",\"introduction\":\"新简介\",\"tags\":\"[\\\"tag1\\\"]\"}",
         "code=0"),
        ("TC-PE-002", "/picture/edit", "安全测试", 1, "编辑_XSS标签", "testuser01 Cookie",
         "POST /picture/edit\nBody: {\"id\":1,\"tags\":\"[\\\"<script>alert(1)</script>\\\"]\"}",
         "code=0, 渲染时被转义"),
        ("TC-PE-003", "/picture/edit", "功能测试", 2, "编辑_不存在的id", "testuser01 Cookie",
         "POST /picture/edit\nBody: {\"id\":999999,\"name\":\"x\"}",
         "code=40001"),
        # ---- /picture/edit/batch ----
        ("TC-PEB-001", "/picture/edit/batch", "功能测试", 1, "批量编辑_3张分类", "testuser01 Cookie",
         "POST /picture/edit/batch\nBody: {\"ids\":[1,2,3],\"category\":\"风景\"}",
         "code=0, 3张图category被更新"),
        ("TC-PEB-002", "/picture/edit/batch", "功能测试", 2, "批量编辑_空数组", "testuser01 Cookie",
         "POST /picture/edit/batch\nBody: {\"ids\":[]}",
         "code=40001"),
        # ---- /picture/update ----
        ("TC-PUd-001", "/picture/update", "功能测试", 1, "管理员更新图片", "admin Cookie",
         "POST /picture/update\nBody: {\"id\":1,\"name\":\"管理员修改\"}",
         "code=0"),
        ("TC-PUd-002", "/picture/update", "安全测试", 1, "更新_非管理员越权", "testuser01 Cookie",
         "POST /picture/update\nBody: {\"id\":1}",
         "code=40300"),
        # ---- /picture/tag_category ----
        ("TC-PTC-001", "/picture/tag_category", "功能测试", 1, "获取标签分类聚合", "testuser01 Cookie",
         "GET /picture/tag_category",
         "返回 tagList + categoryList"),
        # ---- /picture/review ----
        ("TC-PR-001", "/picture/review", "功能测试", 1, "管理员审核通过", "admin Cookie",
         "POST /picture/review\nBody: {\"id\":1,\"reviewStatus\":1,\"reviewMessage\":\"ok\"}",
         "code=0, reviewStatus=1"),
        ("TC-PR-002", "/picture/review", "功能测试", 1, "管理员审核拒绝", "admin Cookie",
         "POST /picture/review\nBody: {\"id\":1,\"reviewStatus\":2,\"reviewMessage\":\"违规\"}",
         "code=0, reviewStatus=2"),
        ("TC-PR-003", "/picture/review", "安全测试", 1, "审核_普通用户越权", "testuser01 Cookie",
         "POST /picture/review\nBody: {\"id\":1}",
         "code=40300"),
        # ---- /picture/search/picture ----
        ("TC-PSP-001", "/picture/search/picture", "功能测试", 1, "关键字搜图", "testuser01 Cookie",
         "POST /picture/search/picture\nBody: {\"text\":\"猫\"}",
         "返回含'猫'的图片列表"),
        ("TC-PSP-002", "/picture/search/picture", "功能测试", 2, "搜图_无结果", "testuser01 Cookie",
         "POST /picture/search/picture\nBody: {\"text\":\"xyz999\"}",
         "返回空列表"),
        ("TC-PSP-003", "/picture/search/picture", "功能测试", 2, "搜图_空文本", "testuser01 Cookie",
         "POST /picture/search/picture\nBody: {\"text\":\"\"}",
         "code=40001"),
        # ---- /picture/search/color ----
        ("TC-PSC-001", "/picture/search/color", "功能测试", 1, "颜色搜图_蓝色", "testuser01 Cookie",
         "POST /picture/search/color\nBody: {\"picColor\":\"#0000FF\"}",
         "返回主色为蓝的图片"),
        ("TC-PSC-002", "/picture/search/color", "功能测试", 2, "颜色搜图_无效颜色值", "testuser01 Cookie",
         "POST /picture/search/color\nBody: {\"picColor\":\"notacolor\"}",
         "code=40001"),
        # ---- /picture/out_painting ----
        ("TC-POP-001", "/picture/out_painting/create_task", "功能测试", 1, "AI扩图_创建任务", "testuser01 Cookie",
         "POST /picture/out_painting/create_task\nBody: {\"pictureId\":1}",
         "code=0, outputTaskId 非空"),
        ("TC-POG-001", "/picture/out_painting/get_task", "功能测试", 1, "AI扩图_查询任务", "有任务",
         "GET /picture/out_painting/get_task?taskId=xxx",
         "返回任务状态"),
        # ---- /picture/proxy/editor ----
        ("TC-PPE-001", "/picture/proxy/editor", "功能测试", 1, "编辑代理_跨域URL", "testuser01 Cookie",
         "GET /picture/proxy/editor?url=https://example.com/1.jpg",
         "返回图片二进制流"),
        ("TC-PPE-002", "/picture/proxy/editor", "安全测试", 1, "编辑代理_内网URL_SSRF", "testuser01 Cookie",
         "GET /picture/proxy/editor?url=http://192.168.1.1/admin",
         "code=40001 拒绝内网"),
        # ---- 性能 ----
        ("TC-PERF-001", "/picture/upload", "性能测试", 2, "上传_20并发", "JMeter",
         "POST /file/upload 20线程 5s内启动\nBody: multipart file=2MB_jpg",
         "P95 < 3s, 错误率 < 5%"),
        # ---- 接口 ----
        ("TC-API-001", "/picture/upload", "接口测试", 2, "上传_缺multipart边界", "testuser01 Cookie",
         "POST /file/upload\nHeader: Content-Type: multipart/form-data (缺 boundary)\nBody: raw 乱码",
         "400 或 415"),
        ("TC-API-002", "/picture/edit", "接口测试", 2, "编辑_id类型错误", "testuser01 Cookie",
         "POST /picture/edit\nBody: {\"id\":\"abc\",\"name\":\"x\"}",
         "code=40001"),
    ],

    "李坤纬": [
        # ---- /space/list/level ----
        ("TC-SPL-001", "/space/list/level", "功能测试", 1, "获取空间等级列表", "无",
         "GET /space/list/level",
         "code=0, 含普通版/专业版/旗舰版"),
        # ---- /space/add ----
        ("TC-SPA-001", "/space/add", "功能测试", 1, "创建空间_普通版", "testuser01 Cookie",
         "POST /space/add\nBody: {\"spaceName\":\"我的空间\",\"spaceLevel\":0}",
         "code=0, spaceId 存在"),
        ("TC-SPA-002", "/space/add", "功能测试", 2, "创建空间_名称为空", "testuser01 Cookie",
         "POST /space/add\nBody: {\"spaceName\":\"\",\"spaceLevel\":0}",
         "code=40001"),
        ("TC-SPA-003", "/space/add", "功能测试", 2, "创建空间_名称超长50", "testuser01 Cookie",
         "POST /space/add\nBody: {\"spaceName\":\"x\"*50,\"spaceLevel\":0}",
         "code=40001"),
        ("TC-SPA-004", "/space/add", "安全测试", 1, "创建空间_未登录", "无 Cookie",
         "POST /space/add\nBody: {\"spaceName\":\"x\",\"spaceLevel\":0}",
         "code=40100"),
        # ---- /space/save ----
        ("TC-SPS-001", "/space/save", "功能测试", 1, "保存空间", "testuser01 Cookie",
         "POST /space/save\nBody: {\"spaceName\":\"x\"}",
         "code=0"),
        # ---- /space/get ----
        ("TC-SPG-001", "/space/get", "功能测试", 1, "按id查询空间", "testuser01 Cookie",
         "GET /space/get?id=1",
         "code=0"),
        ("TC-SPG-002", "/space/get", "功能测试", 2, "查询空间_不存在id", "testuser01 Cookie",
         "GET /space/get?id=999999",
         "code=40001"),
        # ---- /space/get/vo ----
        ("TC-SPGV-001", "/space/get/vo", "功能测试", 1, "按id查询空间VO", "testuser01 Cookie",
         "GET /space/get/vo?id=1",
         "code=0, SpaceVO"),
        # ---- /space/list/page ----
        ("TC-SPLP-001", "/space/list/page", "功能测试", 1, "分页查询空间", "testuser01 Cookie",
         "POST /space/list/page\nBody: {\"current\":1,\"pageSize\":10}",
         "code=0"),
        ("TC-SPLP-002", "/space/list/page", "功能测试", 2, "分页_超大页", "testuser01 Cookie",
         "POST /space/list/page\nBody: {\"current\":1,\"pageSize\":10000}",
         "被限制或code=40001"),
        # ---- /space/list/page/vo ----
        ("TC-SPLPV-001", "/space/list/page/vo", "功能测试", 1, "分页查询空间VO", "testuser01 Cookie",
         "POST /space/list/page/vo\nBody: {\"current\":1,\"pageSize\":10}",
         "code=0"),
        # ---- /space/edit ----
        ("TC-SPE-001", "/space/edit", "功能测试", 1, "编辑空间", "testuser01 Cookie",
         "POST /space/edit\nBody: {\"id\":1,\"spaceName\":\"新名\"}",
         "code=0"),
        ("TC-SPE-002", "/space/edit", "安全测试", 1, "编辑空间_越权", "testuser02 Cookie",
         "POST /space/edit\nBody: {\"id\":1,\"spaceName\":\"x\"}",
         "code=40300"),
        # ---- /space/delete ----
        ("TC-SPD-001", "/space/delete", "功能测试", 1, "删除空间", "testuser01 Cookie",
         "POST /space/delete\nBody: {\"id\":1}",
         "code=0, isDelete=1"),
        ("TC-SPD-002", "/space/delete", "安全测试", 1, "删除空间_越权", "testuser02 Cookie",
         "POST /space/delete\nBody: {\"id\":1}",
         "code=40300"),
        # ---- /space/update ----
        ("TC-SPU-001", "/space/update", "功能测试", 1, "更新空间", "testuser01 Cookie",
         "POST /space/update\nBody: {\"id\":1,\"spaceName\":\"新名\"}",
         "code=0"),
        # ---- /spaceUser/add ----
        ("TC-SUA-001", "/spaceUser/add", "功能测试", 1, "添加空间成员", "admin Cookie",
         "POST /spaceUser/add\nBody: {\"spaceId\":1,\"userId\":2,\"spaceRole\":\"viewer\"}",
         "code=0"),
        ("TC-SUA-002", "/spaceUser/add", "功能测试", 2, "添加成员_重复", "admin Cookie",
         "POST /spaceUser/add\nBody: {\"spaceId\":1,\"userId\":2,\"spaceRole\":\"viewer\"}",
         "code=40001"),
        ("TC-SUA-003", "/spaceUser/add", "安全测试", 1, "添加成员_未登录", "无 Cookie",
         "POST /spaceUser/add\nBody: {\"spaceId\":1,\"userId\":2}",
         "code=40100"),
        # ---- /spaceUser/get ----
        ("TC-SUG-001", "/spaceUser/get", "功能测试", 1, "查询空间成员", "admin Cookie",
         "POST /spaceUser/get\nBody: {\"spaceId\":1,\"userId\":2}",
         "返回 SpaceUser"),
        # ---- /spaceUser/list ----
        ("TC-SUL-001", "/spaceUser/list", "功能测试", 1, "查询空间成员列表", "admin Cookie",
         "POST /spaceUser/list\nBody: {\"spaceId\":1}",
         "返回成员列表"),
        # ---- /spaceUser/list/my ----
        ("TC-SULM-001", "/spaceUser/list/my", "功能测试", 1, "我加入的空间列表", "testuser01 Cookie",
         "POST /spaceUser/list/my",
         "返回 testuser01 加入的空间列表"),
        # ---- /spaceUser/edit ----
        ("TC-SUE-001", "/spaceUser/edit", "功能测试", 1, "编辑空间成员角色", "admin Cookie",
         "POST /spaceUser/edit\nBody: {\"id\":1,\"spaceRole\":\"editor\"}",
         "code=0"),
        # ---- /spaceUser/delete ----
        ("TC-SUD-001", "/spaceUser/delete", "功能测试", 1, "删除空间成员", "admin Cookie",
         "POST /spaceUser/delete\nBody: {\"id\":1}",
         "code=0, isDelete=1"),
        ("TC-SUD-002", "/spaceUser/delete", "安全测试", 1, "删除成员_越权", "testuser02 Cookie",
         "POST /spaceUser/delete\nBody: {\"id\":1}",
         "code=40300"),
        # ---- /space/analyze/usage ----
        ("TC-SA-001", "/space/analyze/usage", "功能测试", 1, "空间使用情况分析", "testuser01 Cookie",
         "POST /space/analyze/usage\nBody: {\"spaceId\":1}",
         "返回 usedSize, maxSize, usagePercent"),
        # ---- /space/analyze/category ----
        ("TC-SAC-001", "/space/analyze/category", "功能测试", 1, "空间分类分析", "testuser01 Cookie",
         "POST /space/analyze/category\nBody: {\"spaceId\":1}",
         "返回分类统计"),
        # ---- /space/analyze/tag ----
        ("TC-SAT-001", "/space/analyze/tag", "功能测试", 1, "空间标签分析", "testuser01 Cookie",
         "POST /space/analyze/tag\nBody: {\"spaceId\":1}",
         "返回标签统计"),
        # ---- /space/analyze/size ----
        ("TC-SASZ-001", "/space/analyze/size", "功能测试", 1, "空间大小分析", "testuser01 Cookie",
         "POST /space/analyze/size\nBody: {\"spaceId\":1}",
         "返回 size 区间分布"),
        # ---- /space/analyze/user ----
        ("TC-SAU-001", "/space/analyze/user", "功能测试", 1, "空间用户分析", "testuser01 Cookie",
         "POST /space/analyze/user\nBody: {\"spaceId\":1}",
         "返回用户上传统计"),
        # ---- /space/analyze/rank ----
        ("TC-SAR-001", "/space/analyze/rank", "功能测试", 1, "空间排行", "testuser01 Cookie",
         "POST /space/analyze/rank",
         "返回空间使用排行"),
        # ---- 性能 ----
        ("TC-PERF-001", "/space/analyze/usage", "性能测试", 2, "空间分析_20并发", "JMeter",
         "POST /space/analyze/usage 20线程 5s内启动\nBody: {\"spaceId\":1}",
         "P95 < 1s, 错误率 0%"),
        # ---- 接口 ----
        ("TC-API-001", "/space/analyze/usage", "接口测试", 2, "分析_缺spaceId", "testuser01 Cookie",
         "POST /space/analyze/usage\nBody: {}",
         "code=40001"),
    ],

    "林景彬": [
        # ---- /file/test/upload ----
        ("TC-FT-001", "/file/test/upload", "功能测试", 1, "test_upload_正常jpg", "testuser01 Cookie",
         "POST /file/test/upload\nBody: multipart file=1MB_jpg",
         "code=0"),
        # ---- /file/upload ----
        ("TC-FL-001", "/file/upload", "功能测试", 1, "本地上传_正常jpg", "testuser01 Cookie",
         "POST /file/upload\nBody: multipart file=2MB_jpg",
         "code=0"),
        ("TC-FL-002", "/file/upload", "功能测试", 1, "本地上传_png", "testuser01 Cookie",
         "POST /file/upload\nBody: multipart file=1MB_png",
         "code=0"),
        ("TC-FL-003", "/file/upload", "功能测试", 2, "本地上传_超2MB", "testuser01 Cookie",
         "POST /file/upload\nBody: multipart file=5MB_jpg",
         "code=40001"),
        ("TC-FL-004", "/file/upload", "功能测试", 2, "本地上传_非图片", "testuser01 Cookie",
         "POST /file/upload\nBody: multipart file=test.pdf",
         "code=40001"),
        ("TC-FL-005", "/file/upload", "功能测试", 2, "本地上传_空文件", "testuser01 Cookie",
         "POST /file/upload\nBody: multipart file=0byte.jpg",
         "code=40001"),
        ("TC-FL-006", "/file/upload", "安全测试", 1, "上传_伪JSP木马", "testuser01 Cookie",
         "POST /file/upload\nBody: multipart file=1.jpg (内容是 <?php system($_GET['c']);?>)",
         "code=40001"),
        ("TC-FL-007", "/file/upload", "安全测试", 1, "上传_双扩展名绕过", "testuser01 Cookie",
         "POST /file/upload\nBody: multipart file=1.jpg.php",
         "code=40001"),
        ("TC-FL-008", "/file/upload", "功能测试", 2, "本地上传_未登录", "无 Cookie",
         "POST /file/upload\nBody: multipart file=1MB_jpg",
         "code=40100"),
        # ---- /file/upload/avatar ----
        ("TC-FA-001", "/file/upload/avatar", "功能测试", 1, "头像上传_jpg", "testuser01 Cookie",
         "POST /file/upload/avatar\nBody: multipart file=512KB_jpg",
         "code=0, userAvatar 已更新"),
        ("TC-FA-002", "/file/upload/avatar", "功能测试", 1, "头像上传_png", "testuser01 Cookie",
         "POST /file/upload/avatar\nBody: multipart file=512KB_png",
         "code=0"),
        ("TC-FA-003", "/file/upload/avatar", "功能测试", 2, "头像上传_超2MB", "testuser01 Cookie",
         "POST /file/upload/avatar\nBody: multipart file=5MB_jpg",
         "code=40001"),
        ("TC-FA-004", "/file/upload/avatar", "功能测试", 2, "头像上传_非图片", "testuser01 Cookie",
         "POST /file/upload/avatar\nBody: multipart file=test.pdf",
         "code=40001"),
        ("TC-FA-005", "/file/upload/avatar", "功能测试", 2, "头像上传_空文件", "testuser01 Cookie",
         "POST /file/upload/avatar\nBody: multipart file=0byte.jpg",
         "code=40001"),
        # ---- /user/get/login (林景彬也测这个) ----
        ("TC-GC-001", "/user/get/login", "功能测试", 1, "获取当前用户_已登录", "testuser01 Cookie",
         "GET /user/get/login\nCookie: xxx",
         "code=0, 含 userAvatar"),
        ("TC-GC-002", "/user/get/login", "功能测试", 1, "获取当前用户_未登录", "无 Cookie",
         "GET /user/get/login",
         "code=40100"),
        ("TC-GC-003", "/user/get/login", "功能测试", 2, "获取当前用户_Cookie过期", "过期 Cookie",
         "GET /user/get/login\nCookie: 过期的JWT",
         "code=40100"),
        ("TC-GC-004", "/user/get/login", "安全测试", 1, "获取当前用户_Cookie伪造", "伪造 Cookie",
         "GET /user/get/login\nCookie: userId=999999 的伪造Cookie",
         "code=40100"),
        # ---- /wx/mp/portal ----
        ("TC-WP-001", "/wx/mp/portal", "功能测试", 1, "公众号门户_签名验证", "微信服务器GET",
         "GET /wx/mp/portal?signature=xxx&timestamp=xxx&nonce=xxx&echostr=xxx",
         "返回 echostr 明文"),
        ("TC-WP-002", "/wx/mp/portal", "功能测试", 1, "公众号门户_消息处理", "微信服务器POST",
         "POST /wx/mp/portal\nBody: <xml>...</xml>",
         "code=0, 响应XML"),
        # ---- /wx/mp/menu/create ----
        ("TC-WM-001", "/wx/mp/menu/create", "功能测试", 1, "创建公众号菜单", "admin Cookie",
         "POST /wx/mp/menu/create\nBody: {\"button\":[{\"name\":\"测试\",\"type\":\"click\",\"key\":\"test\"}]}",
         "code=0"),
        ("TC-WM-002", "/wx/mp/menu/create", "功能测试", 2, "创建菜单_无菜单数据", "admin Cookie",
         "POST /wx/mp/menu/create\nBody: {}",
         "code=40001"),
        ("TC-WM-003", "/wx/mp/menu/create", "安全测试", 1, "创建菜单_未登录", "无 Cookie",
         "POST /wx/mp/menu/create\nBody: {...}",
         "code=40100"),
        # ---- 性能 ----
        ("TC-PERF-001", "/file/upload", "性能测试", 2, "文件上传_50并发", "JMeter",
         "POST /file/upload 50线程 5s内启动\nBody: multipart file=1MB_jpg",
         "P95 < 2s, 错误率 < 5%"),
        # ---- 接口 ----
        ("TC-API-001", "/file/upload", "接口测试", 2, "upload_缺multipart边界", "testuser01 Cookie",
         "POST /file/upload\nHeader: Content-Type: multipart/form-data (缺 boundary)\nBody: raw 乱码",
         "400 或 415"),
    ],
}

# ============================================================
# 2. 生成 4 份 xlsx + 4 份 Postman JSON
# ============================================================

HEADERS = ["用例编号", "所属产品", "所属模块", "用例类型", "优先级", "用例标题",
           "前置条件", "步骤", "预期结果", "实测结果", "结论", "测试人员", "测试时间"]

TYPE_MAP = {1: "功能测试", 2: "性能测试", 3: "接口测试", 4: "安全测试",
            5: "兼容性测试", 6: "自动化功能测试", 7: "单元测试"}


def make_postman_collection(cases):
    """从用例列表生成可导入的 Postman 集合 JSON"""
    items = []
    for c in cases:
        tc_id, module, test_type, priority, title, precondition, steps, expected = c
        method = "GET"
        url = module
        body = None

        # 从步骤中解析 method 和 body
        if steps.startswith("GET"):
            method = "GET"
        elif steps.startswith("POST"):
            method = "POST"

        # 从步骤中解析 body
        body_lines = []
        in_body = False
        for line in steps.split("\n"):
            line = line.strip()
            if line.startswith("Body:") or line.startswith("Body: "):
                in_body = True
                body_lines.append(line.replace("Body:", "").strip())
            elif in_body and line:
                body_lines.append(line)
        if body_lines:
            raw = " ".join(body_lines)
            # 尝试 parse JSON
            try:
                body = json.loads(raw.replace("\\" , "\\\\"))
            except:
                body = raw

        url_raw = f"http://localhost:8123/api{module}"

        request = {
            "name": f"{tc_id} {title}",
            "request": {
                "method": method,
                "header": [{"key": "Content-Type", "value": "application/json"}],
                "url": {
                    "raw": url_raw,
                    "protocol": "http",
                    "host": ["localhost"],
                    "port": "8123",
                    "path": ["api"] + [p for p in module.strip("/").split("/") if p]
                },
                "description": f"前置条件: {precondition}\n\n步骤:\n{steps}\n\n预期结果:\n{expected}"
            }
        }
        if body:
            request["request"]["body"] = {
                "mode": "raw",
                "raw": json.dumps(body) if isinstance(body, (dict, list)) else str(body),
                "options": {"raw": {"language": "json"}}
            }
        items.append(request)

    return {
        "info": {
            "name": f"StarPicture 测试用例集",
            "description": "可直接导入 Postman 运行的测试集合",
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"
        },
        "item": items
    }


def build_xlsx(member, cases):
    """生成单人 xlsx"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append(HEADERS)

    for c in cases:
        tc_id, module, test_type, priority, title, precondition, steps, expected = c
        row = [tc_id, PRODUCT, module, TYPE_MAP.get(test_type, "功能测试"),
               priority, title, precondition, steps, expected, "", "", member, ""]
        ws.append(row)

    # 设置列宽
    for col_idx in range(1, 14):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 15

    # 设置样式
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=13):
        for cell in row:
            cell.alignment = left
            cell.border = thin_border

    # 保存
    out = BASE / f"{member}_脚本与截图/软件测试测试用例.xlsx"
    wb.save(str(out))
    print(f"  xlsx: {out} ({len(cases)} 条)")

    # 生成 Postman JSON
    collection = make_postman_collection(cases)
    postman_dir = BASE / f"{member}_脚本与截图/功能测试"
    postman_dir.mkdir(parents=True, exist_ok=True)
    postman_file = postman_dir / f"{member}_functional.postman_collection.json"
    with open(str(postman_file), "w", encoding="utf-8") as f:
        json.dump(collection, f, ensure_ascii=False, indent=2)
    print(f"  Postman: {postman_file}")


# ============================================================
# 3. 生成汇总 xlsx
# ============================================================
def build_summary(all_cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例汇总"
    ws.append(HEADERS)

    for member, cases in all_cases.items():
        for c in cases:
            tc_id, module, test_type, priority, title, precondition, steps, expected = c
            row = [tc_id, PRODUCT, module, TYPE_MAP.get(test_type, "功能测试"),
                   priority, title, precondition, steps, expected, "", "", member, ""]
            ws.append(row)

    # 设置列宽
    for col_idx in range(1, 14):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 15

    # 设置样式
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=13):
        for cell in row:
            cell.alignment = left
            cell.border = thin_border

    out = BASE / "StarPicture_测试用例.xlsx"
    wb.save(str(out))
    total = ws.max_row - 1
    print(f"\n汇总: {out} ({total} 条)")


# ============================================================
# 4. 更新评分表
# ============================================================
def update_scoring(all_cases):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from collections import Counter

    scoring_path = BASE / "StarPicture_评分表.xlsx"
    if not scoring_path.exists():
        print("评分表不存在，跳过")
        return

    wb = load_workbook(str(scoring_path))
    ws = wb.active

    # 成员 -> 用例数映射
    member_counts = {}
    for member, cases in all_cases.items():
        member_counts[member] = len(cases)

    # 找到每个成员的行
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if name and name.strip() in member_counts:
            member = name.strip()
            total = member_counts[member]
            # 更新备注（列 24）
            ws.cell(row, 24).value = f"{member} 实际 {total} 条（功能+性能+接口+安全）"

            # 更新各类型用例数（列 9-15）
            type_counter = Counter(c[3] for c in CASES[member])
            ws.cell(row, 9).value = type_counter.get(1, 0)   # 功能
            ws.cell(row, 10).value = type_counter.get(2, 0)  # 性能
            ws.cell(row, 11).value = type_counter.get(3, 0)  # 接口
            ws.cell(row, 12).value = type_counter.get(4, 0)  # 安全
            ws.cell(row, 13).value = type_counter.get(5, 0)  # 兼容
            ws.cell(row, 14).value = type_counter.get(6, 0)  # 自动化
            ws.cell(row, 15).value = type_counter.get(7, 0)  # 单元

    wb.save(str(scoring_path))
    print(f"评分表已更新: {scoring_path}")


# ============================================================
# 主程序
# ============================================================
if __name__ == "__main__":
    print("=== 生成 xlsx 和 Postman JSON ===\n")
    for member, cases in CASES.items():
        print(f"\n{member} ({len(cases)} 条):")
        build_xlsx(member, cases)

    print("\n=== 生成汇总 xlsx ===")
    build_summary(CASES)

    print("\n=== 更新评分表 ===")
    update_scoring(CASES)

    print("\n=== 完成 ===")
    for m, cases in CASES.items():
        print(f"  {m}: {len(cases)} 条")
    print(f"  总计: {sum(len(v) for v in CASES.values())} 条")
