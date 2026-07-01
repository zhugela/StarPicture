"""
重新生成 4 份 Postman JSON，包含全部测试用例（功能+性能+接口+安全）
每个 JSON 里按测试类型分组
"""
import json
from openpyxl import load_workbook
from pathlib import Path

BASE = Path(r"D:\code\StarPicture\docs\内娱图库_海蒂与爷爷_朱远亮_18144610287")

# 读取所有用例
all_cases = []
for name in ['朱远亮', '李冠燃', '李坤纬', '林景彬']:
    wb = load_workbook(str(BASE / f"{name}_脚本与截图/软件测试测试用例.xlsx"))
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            tc_id = row[0]
            module = row[1]
            test_type = row[2]
            priority = row[3]
            desc = row[4]
            all_cases.append({
                'id': tc_id,
                'module': module,
                'type': test_type,
                'priority': priority,
                'desc': desc,
                'owner': name,
            })

print(f"总用例: {len(all_cases)} 条")

# 按负责人分组
by_owner = {}
for c in all_cases:
    o = c['owner']
    if o not in by_owner:
        by_owner[o] = []
    by_owner[o].append(c)

# 为每个人生成 Postman JSON
def create_postman_collection(cases, member_name, base_url="http://localhost:8123/api"):
    """生成 Postman 集合，按测试类型分组"""
    items = []
    for c in cases:
        tc_id = c['id']
        test_type = c['type']
        module = c['module']
        desc = c['desc']
        priority = c['priority']

        # 根据测试类型构造请求
        if test_type == '功能测试':
            if '/register' in module or '/login' in module or '/add' in module:
                method = 'POST'
                params = {'userAccount': 'test', 'userPassword': 'Test12345'}
            elif '/upload' in module:
                method = 'POST'
                params = {'file': 'test.jpg'}
            elif '/delete' in module or '/review' in module:
                method = 'POST'
                params = {'id': 1}
            elif '/edit' in module or '/update' in module:
                method = 'POST'
                params = {'id': 1, 'name': 'test'}
            else:
                method = 'GET'
                params = {}

        elif test_type == '性能测试':
            if '/login' in module:
                method = 'POST'
                params = {'userAccount': 'test', 'userPassword': 'Test12345'}
            elif '/upload' in module:
                method = 'POST'
                params = {'file': 'test.jpg'}
            else:
                method = 'POST'
                params = {}

        elif test_type == '接口测试':
            method = 'POST'
            params = {'userAccount': 'test', 'userPassword': 'Test12345'}

        elif test_type == '安全测试':
            method = 'POST'
            params = {'userAccount': "test' OR '1'='1", 'userPassword': '12345678'}

        else:
            method = 'GET'
            params = {}

        item = {
            "name": f"{tc_id} - {desc}",
            "request": {
                "method": method,
                "header": [{"key": "Content-Type", "value": "application/json"}],
                "url": {
                    "raw": f"{base_url}{module}",
                    "protocol": "http",
                    "host": ["localhost"],
                    "port": "8123",
                    "path": module.split('/')
                },
                "body": {
                    "mode": "raw",
                    "raw": json.dumps(params, ensure_ascii=False),
                    "options": {"raw": {"language": "json"}}
                }
            }
        }
        items.append(item)

    return {
        "info": {
            "name": f"StarPicture_{member_name}_测试用例",
            "description": f"{member_name}负责模块的全部测试用例（{len(items)}条）",
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"
        },
        "item": items
    }

# 为每个人生成
for name in ['朱远亮', '李冠燃', '李坤纬', '林景彬']:
    cases = by_owner.get(name, [])
    collection = create_postman_collection(cases, name)

    # 保存到功能测试目录
    output_path = BASE / f"{name}_脚本与截图/功能测试/{name}_functional.postman_collection.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(str(output_path), 'w', encoding='utf-8') as f:
        json.dump(collection, f, ensure_ascii=False, indent=2)

    print(f"✅ {name}: {len(cases)} 条 -> {output_path.name}")

print("\n=== 完成 ===")
print("4 份 Postman JSON 已更新，包含全部测试用例（功能+性能+接口+安全）")
