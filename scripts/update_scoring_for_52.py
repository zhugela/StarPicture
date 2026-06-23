"""
更新现有评分表的真实数（按新 52 条用例）
不动模板结构，只改 4 行的用例数量和贡献度
"""
from openpyxl import load_workbook
from pathlib import Path

BASE = Path("D:/code/StarPicture/docs/test")
dst = BASE / "StarPicture_评分表.xlsx"

wb = load_workbook(dst)
ws = wb.active

# 数据：4 人，每人贡献度 + 7 个类型用例数 + 总数
NEW_DATA = [
    # (贡献度, 功能, 性能, 接口, 安全, 自动化, 单元, 兼容, 总数)
    (1.1, 10, 1, 1, 1, 0, 0, 0, 13),  # 朱远亮（组长，贡献度 1.1）
    (1.0, 10, 1, 1, 1, 0, 0, 0, 13),  # 李冠燃
    (1.0, 10, 1, 1, 1, 0, 0, 0, 13),  # 李坤纬
    (1.0, 10, 1, 1, 1, 0, 0, 0, 13),  # 林景彬
]
NAMES = ["朱远亮", "李冠燃", "李坤纬", "林景彬"]

for i, name in enumerate(NAMES):
    r = 3 + i
    gx, gong, xing, jie, an, zi, dan, rong, total = NEW_DATA[i]
    # 名字
    if ws.cell(r, 1).value:
        ws.cell(r, 1).value = name
    ws.cell(r, 2).value = name
    ws.cell(r, 3).value = gx
    ws.cell(r, 10).value = gong
    ws.cell(r, 11).value = xing
    ws.cell(r, 12).value = jie
    ws.cell(r, 13).value = an
    ws.cell(r, 14).value = zi
    ws.cell(r, 15).value = dan
    ws.cell(r, 16).value = rong
    ws.cell(r, 24).value = f"{name} 实际 {total} 条"

wb.save(dst)
print(f"已更新: {dst}")

# 验证
wb = load_workbook(dst)
ws = wb.active
print("\n=== 评分表验证 ===")
for r in range(3, 7):
    name = ws.cell(r,1).value or ws.cell(r,2).value or ''
    gx = ws.cell(r,3).value
    cells = [ws.cell(r,c).value for c in range(10,17)]
    note = ws.cell(r,24).value or ''
    print(f"  {name:<8} 贡献度={gx} | 功能{cells[0]} 性能{cells[1]} 接口{cells[2]} 安全{cells[3]} 自动{cells[4]} 单元{cells[5]} 兼容{cells[6]} | {note}")
