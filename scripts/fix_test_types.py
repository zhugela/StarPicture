"""
修复 xlsx 里的测试类型（功能/性能/接口/安全）
基于用例编号的前缀来判断：
- TC-SEC-xxx → 安全测试
- TC-PERF-xxx → 性能测试
- TC-API-xxx → 接口测试
- 其他 → 功能测试

注意：需要先关闭 Excel 才能修改 xlsx
"""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path
import shutil
import os

BASE = Path(r"D:\code\StarPicture\docs\内娱图库_海蒂与爷爷_朱远亮_18144610287")

def get_test_type(tc_id):
    """根据用例编号判断测试类型"""
    if tc_id.startswith('TC-SEC'):
        return '安全测试'
    elif tc_id.startswith('TC-PERF'):
        return '性能测试'
    elif tc_id.startswith('TC-API'):
        return '接口测试'
    else:
        return '功能测试'

def fix_xlsx_in_place(filepath, member_name):
    """直接修改 xlsx 文件"""
    wb = load_workbook(str(filepath))
    ws = wb.active

    fixed_count = 0
    for row in range(2, ws.max_row + 1):
        tc_id = ws.cell(row, 1).value
        if tc_id:
            old_type = ws.cell(row, 3).value
            new_type = get_test_type(tc_id)
            if old_type != new_type:
                ws.cell(row, 3).value = new_type
                fixed_count += 1

    try:
        wb.save(str(filepath))
        print(f"  ✅ {member_name}: 修复 {fixed_count} 条类型")
        return True
    except PermissionError:
        print(f"  ❌ {member_name}: 文件被锁（请关闭 Excel）")
        return False

def fix_xlsx_copy(filepath, member_name):
    """复制后修改（如果原文件被锁）"""
    temp_path = str(filepath) + '.temp'
    try:
        shutil.copy2(str(filepath), temp_path)
        wb = load_workbook(temp_path)
        ws = wb.active

        fixed_count = 0
        for row in range(2, ws.max_row + 1):
            tc_id = ws.cell(row, 1).value
            if tc_id:
                old_type = ws.cell(row, 3).value
                new_type = get_test_type(tc_id)
                if old_type != new_type:
                    ws.cell(row, 3).value = new_type
                    fixed_count += 1

        wb.save(temp_path)
        shutil.move(temp_path, str(filepath))
        print(f"  ✅ {member_name}: 修复 {fixed_count} 条类型（使用复制模式）")
        return True
    except Exception as e:
        print(f"  ❌ {member_name}: 错误 - {e}")
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return False

# 主逻辑
print("=== 修复 xlsx 测试类型 ===")
print("（需要先关闭 Excel）")
print()

# 尝试直接修改，如果失败则用复制模式
for name in ['朱远亮', '李冠燃', '李坤纬', '林景彬']:
    filepath = BASE / f"{name}_脚本与截图/软件测试测试用例.xlsx"
    if filepath.exists():
        success = fix_xlsx_in_place(filepath, name)
        if not success:
            fix_xlsx_copy(filepath, name)

# 修复汇总 xlsx
print("\n=== 修复汇总 xlsx ===")
summary_path = BASE / 'StarPicture_测试用例.xlsx'
if summary_path.exists():
    try:
        wb = load_workbook(str(summary_path))
        ws = wb['测试用例汇总']
        fixed = 0
        for row in range(2, ws.max_row + 1):
            tc_id = ws.cell(row, 1).value
            if tc_id:
                old_type = ws.cell(row, 3).value
                new_type = get_test_type(tc_id)
                if old_type != new_type:
                    ws.cell(row, 3).value = new_type
                    fixed += 1
        wb.save(str(summary_path))
        print(f"  ✅ 汇总 xlsx: 修复 {fixed} 条类型")
    except Exception as e:
        print(f"  ❌ 汇总 xlsx: 错误 - {e}")

# 验证
print("\n=== 验证修复结果 ===")
try:
    from collections import Counter
    wb = load_workbook(str(BASE / 'StarPicture_测试用例.xlsx'))
    ws = wb['测试用例汇总']
    type_count = Counter()
    owner_type_count = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            tc_id = row[0]
            owner = row[11]
            test_type = row[2]
            type_count[test_type] += 1
            if owner not in owner_type_count:
                owner_type_count[owner] = Counter()
            owner_type_count[owner][test_type] += 1

    print(f"汇总 xlsx 类型分布: {dict(type_count)}")
    for owner, counts in owner_type_count.items():
        total = sum(counts.values())
        print(f"  {owner}: {dict(counts)}, 总计 {total}")
except Exception as e:
    print(f"  ❌ 验证失败: {e}")
    print("  请关闭 Excel 后重试")
