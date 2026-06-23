"""
按新 52 条用例重新生成：
1. 总用例汇总 xlsx
2. 更新评分表（按真实数填）
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil
from pathlib import Path

BASE = Path("D:/code/StarPicture/docs/test")
PRODUCT = "内娱图库StarPicture"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="305496")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(border_style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["用例编号", "所属产品", "所属模块", "用例类型", "优先级", "用例标题",
           "前置条件", "步骤", "预期结果", "实测结果", "结论", "测试人员", "测试时间"]

# ============ 1. 重新生成总用例汇总 ============
wb = Workbook()
ws = wb.active
ws.title = "测试用例汇总"
ws.append(HEADERS)
for c_idx in range(1, len(HEADERS)+1):
    cell = ws.cell(row=1, column=c_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

total = 0
for m in ["朱远亮", "李冠燃", "李坤纬", "林景彬"]:
    p = BASE / f"{m}_脚本与截图/软件测试测试用例.xlsx"
    if not p.exists():
        continue
    wbm = load_workbook(p, data_only=True)
    wsm = wbm["测试用例"]
    for row in wsm.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        row = list(row)
        row[11] = m
        ws.append(row)
        total += 1

for r in range(2, ws.max_row+1):
    for c_idx in range(1, len(HEADERS)+1):
        cell = ws.cell(row=r, column=c_idx)
        cell.alignment = left
        cell.border = border
        if c_idx in (3, 4, 5):
            cell.alignment = center
        if c_idx == 6:
            cell.font = Font(bold=True)

col_widths = [14, 18, 16, 14, 6, 30, 28, 38, 32, 14, 8, 10, 12]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 30
for r in range(2, ws.max_row+1):
    ws.row_dimensions[r].height = 80
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:M{ws.max_row}"

# 统计 sheet
ws2 = wb.create_sheet("汇总统计")
ws2.append(["项目", "数据"])
ws2.append([])
ws2.append(["按人员", "用例数"])
for m in ["朱远亮", "李冠燃", "李坤纬", "林景彬"]:
    count = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r and r[11] == m)
    ws2.append([m, count])
ws2.append(["合计", total])

ws2.append([])
ws2.append(["按测试类型", "用例数"])
type_count = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[3]:
        type_count[r[3]] = type_count.get(r[3], 0) + 1
for k, v in sorted(type_count.items(), key=lambda x: -x[1]):
    ws2.append([k, v])

for r in range(1, ws2.max_row+1):
    for c_idx in range(1, 3):
        cell = ws2.cell(row=r, column=c_idx)
        cell.alignment = center
        cell.border = border
        if r == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 12

out_summary = BASE / "StarPicture_测试用例.xlsx"
wb.save(out_summary)
print(f"已生成汇总: {out_summary}, 共 {total} 条")

# ============ 2. 更新评分表 ============
# 从模板原样复制（保留颜色和结构）
src_template = r"D:/cxdownload/商品管理系统_第六组_韩东晓_13232956005/商品管理系统_第六组_韩东晓_13232956005/第六组评分表.xlsx"
dst_scoring = BASE / "StarPicture_评分表.xlsx"
shutil.copy(src_template, dst_scoring)
print(f"已复制模板到: {dst_scoring}")

wbs = load_workbook(dst_scoring)
wss = wbs.active

# 找出数据行并改名
NAME_MAP = {
    "韩东晓": ("朱远亮", 1.1, 10, 1, 1, 1, 0, 0, 0, 13),
    "陈雨腾": ("李冠燃", 1.0, 10, 1, 1, 1, 0, 0, 0, 13),
    "甘松明": ("李坤纬", 1.0, 10, 1, 1, 1, 0, 0, 0, 13),
    "吴泽波": ("林景彬", 1.0, 10, 1, 1, 1, 0, 0, 0, 13),
}

# 删多余行（如果有第 7 行杨文星）
all_data_rows = []
for r in range(3, wss.max_row + 1):
    v = wss.cell(r, 1).value or wss.cell(r, 2).value
    if v and ("韩东晓" in str(v) or "陈雨腾" in str(v) or "甘松明" in str(v) or "吴泽波" in str(v) or "杨文星" in str(v)):
        all_data_rows.append(r)
if len(all_data_rows) > 4:
    for r in reversed(all_data_rows[4:]):
        wss.delete_rows(r)

# 改 4 行
for r in range(3, 7):
    leader = wss.cell(r, 1).value
    member = wss.cell(r, 2).value
    key = leader or member
    if key in NAME_MAP:
        new_name, gx, gong, xing, jie, an, zi, dan, rong, total_cases = NAME_MAP[key]
        if leader == key:
            wss.cell(r, 1).value = new_name
            if member == key:
                wss.cell(r, 2).value = new_name
        else:
            wss.cell(r, 2).value = new_name
        wss.cell(r, 3).value = gx
        # 列 10-16: 功能 性能 接口 安全 自动化 单元 兼容
        wss.cell(r, 10).value = gong
        wss.cell(r, 11).value = xing
        wss.cell(r, 12).value = jie
        wss.cell(r, 13).value = an
        wss.cell(r, 14).value = zi
        wss.cell(r, 15).value = dan
        wss.cell(r, 16).value = rong
        wss.cell(r, 24).value = f"{new_name} 实际 {total_cases} 条"

wbs.save(dst_scoring)
print(f"已保存评分表")

# ============ 3. 验证 ============
print("\n=== 验证 ===")
for r in range(3, 7):
    name = wss.cell(r,1).value or wss.cell(r,2).value or ''
    gx = wss.cell(r,3).value
    cells = [wss.cell(r,c).value for c in range(10,17)]
    note = wss.cell(r,24).value or ''
    print(f"  {name:<8} 贡献度={gx} | 功能{cells[0]} 性能{cells[1]} 接口{cells[2]} 安全{cells[3]} 自动{cells[4]} 单元{cells[5]} 兼容{cells[6]} | {note}")
