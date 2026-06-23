"""
清掉列 9-15 的错误数据，按真实 13 条用例填回
正确数据：每人 10 功能 + 1 性能 + 1 接口 + 1 安全 + 0 自动 + 0 单元 + 0 兼容 = 13
列 16 用 SUM 公式：=SUM(I{r}:O{r})
列 20 的"测试类型数量 20分"公式保留不动（它引用 I3:J3 等）
"""
from openpyxl import load_workbook
from pathlib import Path

BASE = Path("D:/code/StarPicture/docs/test")
scoring = BASE / "StarPicture_评分表.xlsx"

wb = load_workbook(scoring)
ws = wb.active

# 4 人数据：列 9-15
NEW = {
    "朱远亮": (10, 1, 1, 1, 0, 0, 0),
    "李冠燃": (10, 1, 1, 1, 0, 0, 0),
    "李坤纬": (10, 1, 1, 1, 0, 0, 0),
    "林景彬": (10, 1, 1, 1, 0, 0, 0),
}

for r in range(3, 7):
    name = ws.cell(r, 1).value or ws.cell(r, 2).value or ''
    if name in NEW:
        vals = NEW[name]
        # 清掉列 9-15
        for c in range(9, 16):
            ws.cell(r, c).value = None
        # 重填
        for i, v in enumerate(vals):
            ws.cell(r, 9 + i).value = v
        # 列 16 = SUM 公式
        ws.cell(r, 16).value = f"=SUM(I{r}:O{r})"
        # 备注
        ws.cell(r, 24).value = f"{name} 实际 13 条（功能 10 + 性能 1 + 接口 1 + 安全 1）"

wb.save(scoring)
print(f"已修复")

# 验证
print("\n=== 最终验证 ===")
wb = load_workbook(scoring, data_only=False)
ws = wb.active
headers = ["功能", "性能", "接口", "安全", "自动", "单元", "兼容", "合计"]
print(f"  {'姓名':<8}", end=' ')
for h in headers:
    print(f"{h:>4}", end=' ')
print()
for r in range(3, 7):
    name = ws.cell(r,1).value or ws.cell(r,2).value or ''
    print(f"  {name:<8}", end=' ')
    for c in range(9, 16):
        v = ws.cell(r,c).value
        print(f"{str(v):>4}", end=' ')
    total = ws.cell(r, 16).value
    print(f"合计公式={total}")
    print(f"    备注: {ws.cell(r,24).value}")