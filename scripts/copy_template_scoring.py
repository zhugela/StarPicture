"""
复用第六组评分表.xlsx 做模板：
1. 只保留 4 个人的行（删掉杨文星那行）
2. 改 4 行的名字
3. 按实际生成的用例重新填：贡献度、用例类型数量、用例总数
不改色、不改结构
"""
from openpyxl import load_workbook
from copy import copy
import shutil

src = r"D:/cxdownload/商品管理系统_第六组_韩东晓_13232956005/商品管理系统_第六组_韩东晓_13232956005/第六组评分表.xlsx"
dst = r"D:/code/StarPicture/docs/test/评分表.xlsx"

shutil.copy(src, dst)
wb = load_workbook(dst)
ws = wb.active

# ===== 1. 找出要删的行 =====
# 数据从第 3 行开始
ROWS_TO_KEEP = 4  # 只保留前 4 行（韩东晓、陈雨腾、甘松明、吴泽波）
data_start = 3
data_end = data_start + ROWS_TO_KEEP - 1  # 6

# 找出所有数据行（值为姓名）
all_data_rows = []
for r in range(data_start, ws.max_row + 1):
    v = ws.cell(r, 1).value or ws.cell(r, 2).value
    if v and ("韩东晓" in str(v) or "陈雨腾" in str(v) or "甘松明" in str(v) or "吴泽波" in str(v) or "杨文星" in str(v) or "朱远亮" in str(v) or "李冠燃" in str(v) or "李坤纬" in str(v) or "林景彬" in str(v)):
        all_data_rows.append(r)

# 删除多余的行（从下往上删，避免索引错位）
if len(all_data_rows) > ROWS_TO_KEEP:
    for r in reversed(all_data_rows[ROWS_TO_KEEP:]):
        ws.delete_rows(r)
        print(f"已删除第 {r} 行")

# ===== 2. 重新计算数据行 =====
# 现在数据行应该是 3, 4, 5, 6
# ===== 3. 改名字 + 填实际数据 =====
# 实际生成的 207 条用例（按人员+类型）：
#   朱远亮(user):   53 功能 + 1 性能 + 2 接口 + 3 安全 + 3 兼容 + 1 自动 + 2 单元 = 65
#   李冠燃(picture): 58 功能 + 2 性能 + 2 接口 + 3 安全 + 3 兼容 + 2 自动 + 2 单元 = 72
#   李坤纬(space):  32 功能 + 1 性能 + 1 接口 + 2 安全 + 1 兼容 + 1 自动 + 1 单元 = 39
#   林景彬(file+wxMp): 19 功能 + 1 性能 + 2 接口 + 3 安全 + 2 兼容 + 2 自动 + 2 单元 = 31

NAME_MAP = {
    "韩东晓": ("朱远亮", 1.1, 53, 1, 2, 3, 1, 2, 3, 65),  # 组长，贡献度 1.1
    "陈雨腾": ("李冠燃", 1.0, 58, 2, 2, 3, 2, 2, 3, 72),
    "甘松明": ("李坤纬", 1.0, 32, 1, 1, 2, 1, 1, 1, 39),
    "吴泽波": ("林景彬", 1.0, 19, 1, 2, 3, 2, 2, 2, 31),
}

# 列：1=组长 2=组员 3=贡献度 4=测试计划 5=测试说明 6=测试报告 7=(空) 8=测试文档分数 9=用例执行完成否
# 10=功能 11=性能 12=接口 13=安全 14=自动 15=单元 16=兼容 17=用例数量(公式) 18=用例数量得分(公式)
# 19=用例质量等级 20=用例质量 21=测试类型数量 22=问题回答等级 23=问题回答 24=备注

for r in range(3, 7):
    leader = ws.cell(r, 1).value
    member = ws.cell(r, 2).value
    key = leader or member
    if key in NAME_MAP:
        new_name, gx, gong, xing, jie, an, zi, dan, rong, total = NAME_MAP[key]
        if leader == key:
            ws.cell(r, 1).value = new_name
            if member == key:
                ws.cell(r, 2).value = new_name
        else:
            ws.cell(r, 2).value = new_name
        # 改贡献度
        ws.cell(r, 3).value = gx
        # 填用例类型数量（按列：功能10 性能11 接口12 安全13 自动化14 单元15 兼容16）
        ws.cell(r, 10).value = gong
        ws.cell(r, 11).value = xing
        ws.cell(r, 12).value = jie
        ws.cell(r, 13).value = an
        ws.cell(r, 14).value = zi
        ws.cell(r, 15).value = dan
        ws.cell(r, 16).value = rong
        # 用例数量 P 列（17）是公式 =SUM(J..O)，自动算
        # 备注
        ws.cell(r, 24).value = f"{new_name} 实际 {total} 条"
        print(f"第 {r} 行 -> {new_name} 贡献度={gx} 总用例={total}")

wb.save(dst)
print("\n完成！")
