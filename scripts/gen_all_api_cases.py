"""
生成全部 API 的测试用例（基于真实 Controller 列表）
58 个 API → 每个 API 3-5 条用例 → 4 人 × 40-60 条 = 约 200 条
不删任何已有文件（截图保留）
"""
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

BASE = Path("D:/code/StarPicture/docs/内娱图库_海蒂与爷爷_朱远亮_18144610287")
PRODUCT = "内娱图库StarPicture"

# =============== 样式 ===============
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="305496")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(border_style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["用例编号", "所属产品", "所属模块", "用例类型", "优先级", "用例标题",
           "前置条件", "步骤", "预期结果", "实测结果", "结论", "测试人员", "测试时间"]

# =============== 朱远亮：user 模块（11 个 API，每个 3 条 = 33 条）==============
zhuyuanliang_cases = [
    # register (3)
    ("TC-UR-001", "/用户注册", "功能测试", 1, "注册_正常", "无账号", "POST /user/register userAccount=Zyl_New01 userPassword=12345678 checkPassword=12345678", "code=0, userId>0"),
    ("TC-UR-002", "/用户注册", "功能测试", 1, "注册_账号已存在", "TC-New01 已存在", "重复注册", "code=40001"),
    ("TC-UR-003", "/用户注册", "功能测试", 1, "注册_密码不一致", "无", "userPassword!=checkPassword", "code=40001"),
    # login (3)
    ("TC-UL-001", "/用户登录", "功能测试", 1, "登录_正常", "testuser01 已存在", "POST /user/login userAccount=testuser01 userPassword=12345678", "code=0, Set-Cookie"),
    ("TC-UL-002", "/用户登录", "功能测试", 1, "登录_密码错误", "testuser01", "wrongPassword", "code=40001"),
    ("TC-UL-003", "/用户登录", "功能测试", 1, "登录_账号不存在", "无", "userAccount=notexist", "code=40001"),
    # get/login (2)
    ("TC-GC-001", "/获取当前用户", "功能测试", 1, "获取_已登录", "已登录 Cookie", "GET /user/get/login", "code=0, 用户信息"),
    ("TC-GC-002", "/获取当前用户", "功能测试", 1, "获取_未登录", "无 Cookie", "GET /user/get/login", "code=40100"),
    # update/my (3)
    ("TC-UM-001", "/更新我的信息", "功能测试", 1, "更新_正常昵称", "已登录", "POST /user/update/my userName='新名'", "code=0"),
    ("TC-UM-002", "/更新我的信息", "功能测试", 1, "更新_昵称为空", "已登录", "userName=''", "code=40001"),
    ("TC-UM-003", "/更新我的信息", "功能测试", 1, "更新_未登录", "无", "POST /user/update/my", "code=40100"),
    # add (admin) (3)
    ("TC-UA-001", "/管理员新增用户", "功能测试", 1, "新增_正常", "admin 已登录", "POST /user/add userAccount=Zyl_Admin01", "code=0"),
    ("TC-UA-002", "/管理员新增用户", "功能测试", 1, "新增_普通用户越权", "testuser01", "POST /user/add", "code=40300"),
    ("TC-UA-003", "/管理员新增用户", "功能测试", 1, "新增_未登录", "无", "POST /user/add", "code=40100"),
    # get (admin) (2)
    ("TC-UG-001", "/管理员获取用户", "功能测试", 1, "按id获取_正常", "admin", "GET /user/get?id=1", "code=0, 用户信息"),
    ("TC-UG-002", "/管理员获取用户", "功能测试", 1, "按id获取_不存在", "admin", "GET /user/get?id=999999", "code=40001"),
    # get/vo (admin) (2)
    ("TC-UV-001", "/管理员获取VO", "功能测试", 1, "获取VO_正常", "admin", "GET /user/get/vo?id=1", "code=0"),
    ("TC-UV-002", "/管理员获取VO", "功能测试", 1, "获取VO_普通用户越权", "testuser01", "GET /user/get/vo?id=1", "code=40300"),
    # delete (admin) (2)
    ("TC-UD-001", "/管理员删除用户", "功能测试", 1, "删除_正常", "admin", "POST /user/delete id=新用户id", "code=0, isDelete=1"),
    ("TC-UD-002", "/管理员删除用户", "功能测试", 1, "删除_普通用户越权", "testuser01", "POST /user/delete", "code=40300"),
    # update (admin) (2)
    ("TC-UU-001", "/管理员更新用户", "功能测试", 1, "更新_正常昵称", "admin", "POST /user/update id=新id userName='改名'", "code=0"),
    ("TC-UU-002", "/管理员更新用户", "功能测试", 1, "更新_普通用户越权", "testuser01", "POST /user/update", "code=40300"),
    # list/page/vo (2)
    ("TC-LP-001", "/管理员分页查询", "功能测试", 1, "分页查询_正常", "admin", "POST /user/list/page/vo current=1 pageSize=10", "code=0, records列表"),
    ("TC-LP-002", "/管理员分页查询", "功能测试", 1, "分页查询_普通用户越权", "testuser01", "POST /user/list/page/vo", "code=40300"),
    # logout (2)
    ("TC-LO-001", "/用户注销", "功能测试", 1, "注销_正常", "已登录", "POST /user/logout", "code=0"),
    ("TC-LO-002", "/用户注销", "功能测试", 1, "注销_未登录", "无", "POST /user/logout", "code=40100"),
    # 性能 (1)
    ("TC-PERF-001", "/用户登录", "性能测试", 2, "登录_50并发", "JMeter", "50线程 POST /user/login", "P95<500ms"),
    # 接口 (1)
    ("TC-API-001", "/用户注册", "接口测试", 2, "注册_Content-Type错误", "无", "Content-Type=application/xml", "415"),
    # 安全 (2)
    ("TC-SEC-001", "/用户登录", "安全测试", 1, "SQL注入_登录", "testuser01", "password=\"' OR 1=1 --\"", "code=40001"),
    ("TC-SEC-002", "/管理员", "安全测试", 1, "越权_普通用户改userRole", "testuser01", "POST /user/update role='admin'", "code=40300"),
]

# =============== 李冠燃：picture 模块（18 个 API，每个 3 条 = 54 条）==============
liguanran_cases = [
    # delete (3)
    ("TC-PD-001", "/图片删除", "功能测试", 1, "删除_自己的图片", "testuser01有图片", "POST /picture/delete id=我的图片", "code=0, isDelete=1"),
    ("TC-PD-002", "/图片删除", "功能测试", 1, "删除_别人的图片_越权", "testuser02", "POST /picture/delete id=别人图片", "code=40300"),
    ("TC-PD-003", "/图片删除", "功能测试", 1, "删除_不存在的图片", "admin", "POST /picture/delete id=999999", "code=40001"),
    # get (2)
    ("TC-PG-001", "/图片查询", "功能测试", 1, "按id查询_正常", "已有图片", "GET /picture/get?id=1", "code=0"),
    ("TC-PG-002", "/图片查询", "功能测试", 1, "按id查询_不存在", "无", "GET /picture/get?id=999999", "code=40001"),
    # get/vo (2)
    ("TC-PV-001", "/图片VO查询", "功能测试", 1, "查询VO_正常", "已有图片", "GET /picture/get/vo?id=1", "code=0, PictureVO"),
    ("TC-PV-002", "/图片VO查询", "功能测试", 1, "查询VO_未登录", "无", "GET /picture/get/vo?id=1", "code=40100"),
    # edit (3)
    ("TC-PE-001", "/图片编辑", "功能测试", 1, "编辑_名称正常", "testuser01有图片", "POST /picture/edit id=1 name='新名'", "code=0"),
    ("TC-PE-002", "/图片编辑", "功能测试", 1, "编辑_越权", "testuser02", "POST /picture/edit id=别人图片", "code=40300"),
    ("TC-PE-003", "/图片编辑", "功能测试", 1, "编辑_不存在", "admin", "POST /picture/edit id=999999", "code=40001"),
    # edit/batch (2)
    ("TC-EB-001", "/图片批量编辑", "功能测试", 1, "批量编辑_正常", "testuser01有多张", "POST /picture/edit/batch ids=[1,2,3] category='风景'", "code=0"),
    ("TC-EB-002", "/图片批量编辑", "功能测试", 1, "批量编辑_越权", "testuser02", "POST /picture/edit/batch ids=[别人图片]", "code=40300"),
    # update (2)
    ("TC-PU-001", "/图片更新", "功能测试", 1, "更新_正常", "admin", "POST /picture/update id=1 name='更新名'", "code=0"),
    ("TC-PU-002", "/图片更新", "功能测试", 1, "更新_普通用户越权", "testuser01", "POST /picture/update", "code=40300"),
    # list/page (3)
    ("TC-PL-001", "/图片分页", "功能测试", 1, "分页查询_默认", "admin", "POST /picture/list/page current=1 pageSize=10", "code=0, records"),
    ("TC-PL-002", "/图片分页", "功能测试", 1, "分页查询_普通用户", "testuser01", "POST /picture/list/page", "code=40100或code=0"),
    ("TC-PL-003", "/图片分页", "功能测试", 1, "分页查询_空页", "admin", "current=9999 pageSize=10", "code=0, records=[]"),
    # list/page/vo (2)
    ("TC-LV-001", "/图片VO分页", "功能测试", 1, "VO分页_正常", "admin", "POST /picture/list/page/vo", "code=0, PictureVO列表"),
    ("TC-LV-002", "/图片VO分页", "功能测试", 1, "VO分页_普通用户", "testuser01", "POST /picture/list/page/vo", "code=40100或code=0"),
    # list/page/vo/cache (1)
    ("TC-VC-001", "/图片缓存分页", "功能测试", 1, "缓存分页_正常", "admin", "POST /picture/list/page/vo/cache", "code=0, 含缓存"),
    # tag_category (1)
    ("TC-TG-001", "/图片标签分类", "功能测试", 1, "获取标签分类", "已有图片", "GET /picture/tag_category", "code=0, tags+categories"),
    # review (3)
    ("TC-RV-001", "/图片审核", "功能测试", 1, "审核_通过", "admin,待审核图片", "POST /picture/review id=1 reviewStatus=1", "code=0"),
    ("TC-RV-002", "/图片审核", "功能测试", 1, "审核_拒绝", "admin", "POST /picture/review id=1 reviewStatus=2", "code=0"),
    ("TC-RV-003", "/图片审核", "功能测试", 1, "审核_普通用户越权", "testuser01", "POST /picture/review", "code=40300"),
    # upload/url (3)
    ("TC-UR-001", "/URL上传", "功能测试", 1, "URL上传_正常", "testuser01", "POST /picture/upload/url fileUrl='https://example.com/1.jpg'", "code=0"),
    ("TC-UR-002", "/URL上传", "功能测试", 1, "URL上传_内网IP_拒绝", "testuser01", "fileUrl='http://127.0.0.1/x.jpg'", "code=40001"),
    ("TC-UR-003", "/URL上传", "功能测试", 1, "URL上传_404", "testuser01", "fileUrl='https://example.com/404.jpg'", "code=40001"),
    # upload/batch (2)
    ("TC-UB-001", "/批量上传", "功能测试", 1, "批量上传_5张", "testuser01", "POST /picture/upload/batch files=[5张jpg]", "code=0, 5条记录"),
    ("TC-UB-002", "/批量上传", "功能测试", 1, "批量上传_超20张", "testuser01", "POST /picture/upload/batch files=[30张]", "code=40001"),
    # search/picture (3)
    ("TC-SP-001", "/关键字搜图", "功能测试", 1, "搜图_正常", "已有图片", "POST /picture/search/picture text='猫'", "code=0, 结果列表"),
    ("TC-SP-002", "/关键字搜图", "功能测试", 1, "搜图_无结果", "无", "text='xyz999'", "code=0, []"),
    ("TC-SP-003", "/关键字搜图", "功能测试", 1, "搜图_SQL注入", "testuser01", "text=\"' OR 1=1 --\"", "code=40001"),
    # search/color (2)
    ("TC-SC-001", "/颜色搜图", "功能测试", 1, "搜图_蓝色", "已有图片", "POST /picture/search/color picColor='#0000FF'", "code=0"),
    ("TC-SC-002", "/颜色搜图", "功能测试", 1, "搜图_无效颜色", "testuser01", "picColor='notacolor'", "code=40001"),
    # out_painting/create_task (2)
    ("TC-OC-001", "/AI扩图", "功能测试", 1, "创建扩图任务", "testuser01有图片", "POST /picture/out_painting/create_task pictureId=1", "code=0, taskId"),
    ("TC-OC-002", "/AI扩图", "功能测试", 1, "任务_未登录", "无", "POST /picture/out_painting/create_task", "code=40100"),
    # out_painting/get_task (1)
    ("TC-OG-001", "/AI扩图查询", "功能测试", 1, "查询任务状态", "已创建任务", "GET /picture/out_painting/get_task?taskId=xxx", "code=0, status"),
    # proxy/editor (2)
    ("TC-PE-001", "/编辑代理", "功能测试", 1, "代理_正常URL", "testuser01", "GET /picture/proxy/editor?url=https://example.com/1.jpg", "返回图片"),
    ("TC-PE-002", "/编辑代理", "功能测试", 1, "代理_内网URL_拒绝", "testuser01", "url='http://192.168.1.1/admin'", "code=40001"),
    # 性能 (2)
    ("TC-PERF-001", "/图片上传", "性能测试", 2, "上传_20并发", "JMeter", "20线程 POST /file/upload", "P95<3s"),
    ("TC-PERF-002", "/图片分页", "性能测试", 2, "分页查询_50并发", "JMeter", "50线程 POST /picture/list/page", "P95<300ms"),
    # 接口 (1)
    ("TC-API-001", "/图片上传", "接口测试", 2, "上传_缺multipart边界", "testuser01", "不完整multipart", "400/415"),
    # 安全 (3)
    ("TC-SEC-001", "/图片上传", "安全测试", 1, "上传_伪PHP木马", "testuser01", "file=1.jpg内容为<?php...?>", "code=40001"),
    ("TC-SEC-002", "/URL上传", "安全测试", 1, "URL上传_file协议", "testuser01", "fileUrl='file:///etc/passwd'", "code=40001"),
    ("TC-SEC-003", "/图片编辑", "安全测试", 1, "编辑_name注入XSS", "testuser01", "name='<script>alert(1)</script>'", "code=0, 转义"),
]

# =============== 李坤纬：space 模块（23 个 API）==============
likunwei_cases = [
    # space: list/level (2)
    ("TC-SL-001", "/空间等级", "功能测试", 1, "获取等级列表_正常", "无", "GET /space/list/level", "code=0, 3个等级"),
    ("TC-SL-002", "/空间等级", "功能测试", 1, "获取等级列表_未登录", "无", "GET /space/list/level 无Cookie", "code=0, 不需要登录"),
    # space: save (2)
    ("TC-SS-001", "/空间保存", "功能测试", 1, "保存_正常", "testuser01", "POST /space/save spaceName='保存测试'", "code=0"),
    ("TC-SS-002", "/空间保存", "功能测试", 1, "保存_名称为空", "testuser01", "spaceName=''", "code=40001"),
    # space: add (3)
    ("TC-SA-001", "/空间新增", "功能测试", 1, "新增_正常", "testuser01", "POST /space/add spaceName='新空间' spaceLevel=0", "code=0"),
    ("TC-SA-002", "/空间新增", "功能测试", 1, "新增_名称超长50", "testuser01", "spaceName=50+'x'", "code=40001"),
    ("TC-SA-003", "/空间新增", "功能测试", 1, "新增_未登录", "无", "POST /space/add", "code=40100"),
    # space: get (2)
    ("TC-SG-001", "/空间获取", "功能测试", 1, "按id获取_正常", "testuser01", "GET /space/get?id=1", "code=0"),
    ("TC-SG-002", "/空间获取", "功能测试", 1, "按id获取_不存在", "无", "GET /space/get?id=999999", "code=40001"),
    # space: get/vo (2)
    ("TC-SV-001", "/空间VO", "功能测试", 1, "获取VO_正常", "testuser01", "GET /space/get/vo?id=1", "code=0"),
    ("TC-SV-002", "/空间VO", "功能测试", 1, "获取VO_未登录", "无", "GET /space/get/vo?id=1", "code=40100"),
    # space: list/page (2)
    ("TC-SP-001", "/空间分页", "功能测试", 1, "分页查询_正常", "testuser01", "POST /space/list/page current=1 pageSize=10", "code=0"),
    ("TC-SP-002", "/空间分页", "功能测试", 1, "分页查询_空页", "testuser01", "current=9999", "code=0, []"),
    # space: list/page/vo (2)
    ("TC-LV-001", "/空间VO分页", "功能测试", 1, "VO分页_正常", "testuser01", "POST /space/list/page/vo", "code=0"),
    ("TC-LV-002", "/空间VO分页", "功能测试", 1, "VO分页_未登录", "无", "POST /space/list/page/vo", "code=40100"),
    # space: edit (2)
    ("TC-SE-001", "/空间编辑", "功能测试", 1, "编辑_正常", "testuser01是创建者", "POST /space/edit id=1 spaceName='改名'", "code=0"),
    ("TC-SE-002", "/空间编辑", "功能测试", 1, "编辑_越权", "testuser02", "POST /space/edit id=别人空间", "code=40300"),
    # space: delete (2)
    ("TC-SD-001", "/空间删除", "功能测试", 1, "删除_正常", "testuser01是创建者", "POST /space/delete id=1", "code=0, isDelete=1"),
    ("TC-SD-002", "/空间删除", "功能测试", 1, "删除_越权", "testuser02", "POST /space/delete id=别人空间", "code=40300"),
    # space: update (2)
    ("TC-SU-001", "/空间更新", "功能测试", 1, "更新_正常", "testuser01是创建者", "POST /space/update id=1 spaceName='更新'", "code=0"),
    ("TC-SU-002", "/空间更新", "功能测试", 1, "更新_越权", "testuser02", "POST /space/update", "code=40300"),
    # spaceUser: manage/add (3)
    ("TC-MA-001", "/成员添加", "功能测试", 1, "管理员添加_正常", "admin", "POST /spaceUser/add spaceId=1 userId=2", "code=0"),
    ("TC-MA-002", "/成员添加", "功能测试", 1, "添加_重复", "admin", "重复添加", "code=40001"),
    ("TC-MA-003", "/成员添加", "功能测试", 1, "添加_普通用户越权", "testuser01", "POST /spaceUser/add", "code=40300"),
    # spaceUser: manage/delete (2)
    ("TC-MD-001", "/成员删除", "功能测试", 1, "删除_正常", "admin", "POST /spaceUser/delete id=1", "code=0"),
    ("TC-MD-002", "/成员删除", "功能测试", 1, "删除_越权", "testuser02", "POST /spaceUser/delete", "code=40300"),
    # spaceUser: manage/get (2)
    ("TC-MG-001", "/成员查询", "功能测试", 1, "查询_正常", "admin", "POST /spaceUser/get spaceId=1 userId=2", "code=0"),
    ("TC-MG-002", "/成员查询", "功能测试", 1, "查询_普通用户", "testuser01", "POST /spaceUser/get", "code=40100或code=0"),
    # spaceUser: manage/list (2)
    ("TC-ML-001", "/成员列表", "功能测试", 1, "列表_正常", "admin", "POST /spaceUser/list spaceId=1", "code=0, 成员列表"),
    ("TC-ML-002", "/成员列表", "功能测试", 1, "列表_未登录", "无", "POST /spaceUser/list", "code=40100"),
    # spaceUser: manage/edit (2)
    ("TC-ME-001", "/成员编辑", "功能测试", 1, "编辑_正常", "admin", "POST /spaceUser/edit id=1 spaceRole='editor'", "code=0"),
    ("TC-ME-002", "/成员编辑", "功能测试", 1, "编辑_越权", "testuser02", "POST /spaceUser/edit", "code=40300"),
    # spaceUser: list/my (2)
    ("TC-LM-001", "/我的空间", "功能测试", 1, "我的空间列表", "testuser01", "POST /spaceUser/list/my", "code=0, 我加入的空间"),
    ("TC-LM-002", "/我的空间", "功能测试", 1, "我的空间_未登录", "无", "POST /spaceUser/list/my", "code=40100"),
    # spaceAnalyze: analyze (7)
    ("TC-AU-001", "/空间用量分析", "功能测试", 1, "用量分析_正常", "testuser01有空间", "POST /space/analyze/usage spaceId=1", "code=0"),
    ("TC-AC-001", "/空间分类分析", "功能测试", 1, "分类分析_正常", "testuser01有空间", "POST /space/analyze/category spaceId=1", "code=0"),
    ("TC-AT-001", "/空间标签分析", "功能测试", 1, "标签分析_正常", "testuser01有空间", "POST /space/analyze/tag spaceId=1", "code=0"),
    ("TC-AZ-001", "/空间大小分析", "功能测试", 1, "大小分析_正常", "testuser01有空间", "POST /space/analyze/size spaceId=1", "code=0"),
    ("TC-AP-001", "/空间用户分析", "功能测试", 1, "用户分析_正常", "testuser01有空间", "POST /space/analyze/user spaceId=1", "code=0"),
    ("TC-AR-001", "/空间排行", "功能测试", 1, "排行_正常", "testuser01", "POST /space/analyze/rank", "code=0"),
    ("TC-AR-002", "/空间排行", "功能测试", 1, "排行_未登录", "无", "POST /space/analyze/rank", "code=40100"),
    # 性能 (2)
    ("TC-PERF-001", "/空间分析", "性能测试", 2, "用量分析_20并发", "JMeter", "20线程 POST /space/analyze/usage", "P95<1s"),
    ("TC-PERF-002", "/空间列表", "性能测试", 2, "空间列表_50并发", "JMeter", "50线程 POST /space/list/page/vo", "P95<500ms"),
    # 接口 (2)
    ("TC-API-001", "/空间新增", "接口测试", 2, "新增_缺Content-Type", "testuser01", "不设置Content-Type", "415"),
    ("TC-API-002", "/成员添加", "接口测试", 2, "添加_缺userId", "admin", "无userId字段", "40001"),
    # 安全 (2)
    ("TC-SEC-001", "/空间管理", "安全测试", 1, "空间名称XSS", "testuser01", "spaceName='<script>alert(1)</script>'", "code=0, 转义"),
    ("TC-SEC-002", "/成员管理", "安全测试", 1, "跨用户添加成员", "testuser02", "POST /spaceUser/add spaceId=别人空间", "code=40300"),
]

# =============== 林景彬：file + wxMp（6 个 API）==============
linjingbin_cases = [
    # file: test/upload (3)
    ("TC-TU-001", "/文件上传", "功能测试", 1, "测试上传_正常jpg", "testuser01", "POST /file/test/upload file=1MB_jpg", "code=0"),
    ("TC-TU-002", "/文件上传", "功能测试", 1, "测试上传_空文件", "testuser01", "file=0byte", "code=40001"),
    ("TC-TU-003", "/文件上传", "功能测试", 1, "测试上传_非图片", "testuser01", "file=test.pdf", "code=40001"),
    # file: upload (3)
    ("TC-FU-001", "/文件上传", "功能测试", 1, "上传_正常png", "testuser01", "POST /file/upload file=1MB_png", "code=0"),
    ("TC-FU-002", "/文件上传", "功能测试", 1, "上传_超过2MB", "testuser01", "file=5MB_jpg", "code=40001"),
    ("TC-FU-003", "/文件上传", "功能测试", 1, "上传_未登录", "无", "POST /file/upload", "code=40100"),
    # file: upload/avatar (3)
    ("TC-AU-001", "/头像上传", "功能测试", 1, "头像_正常", "testuser01", "POST /file/upload/avatar file=300KB_jpg", "code=0, userAvatar更新"),
    ("TC-AU-002", "/头像上传", "功能测试", 1, "头像_超过1MB", "testuser01", "file=1.5MB", "code=40001"),
    ("TC-AU-003", "/头像上传", "功能测试", 1, "头像_非图片", "testuser01", "file=test.pdf", "code=40001"),
    # wxMp: portal GET (3)
    ("TC-WG-001", "/微信门户", "功能测试", 1, "门户_GET_签名验证", "微信GET", "GET /wx/mp/portal signature=xxx timestamp=xxx nonce=xxx echostr=test", "返回echostr"),
    ("TC-WG-002", "/微信门户", "功能测试", 1, "门户_GET_签名错误", "微信GET", "signature=invalid", "返回空或错误"),
    ("TC-WG-003", "/微信门户", "功能测试", 1, "门户_GET_缺参数", "微信GET", "无signature参数", "返回错误"),
    # wxMp: portal POST (3)
    ("TC-WP-001", "/微信门户", "功能测试", 1, "门户_POST_正常文本", "微信POST", "POST /wx/mp/portal body=<xml>...", "code=0, 响应XML"),
    ("TC-WP-002", "/微信门户", "功能测试", 1, "门户_POST_空body", "微信POST", "body=空", "返回错误XML"),
    ("TC-WP-003", "/微信门户", "功能测试", 1, "门户_POST_格式错误", "微信POST", "body=非XML", "返回错误"),
    # wxMp: menu/create (3)
    ("TC-MC-001", "/微信菜单", "功能测试", 1, "创建菜单_正常", "admin", "POST /wx/mp/menu/create menu={button:[...]}", "code=0"),
    ("TC-MC-002", "/微信菜单", "功能测试", 1, "创建菜单_空body", "admin", "body={}", "code=40001"),
    ("TC-MC-003", "/微信菜单", "功能测试", 1, "创建菜单_未登录", "无", "POST /wx/mp/menu/create", "code=40100"),
    # 性能 (2)
    ("TC-PERF-001", "/文件上传", "性能测试", 2, "上传_50并发1MB", "JMeter", "50线程 POST /file/upload", "P95<2s"),
    ("TC-PERF-002", "/头像上传", "性能测试", 2, "头像_100并发", "JMeter", "100线程 POST /file/upload/avatar", "P95<3s"),
    # 接口 (1)
    ("TC-API-001", "/文件上传", "接口测试", 2, "上传_缺multipart边界", "testuser01", "不完整multipart", "400/415"),
    # 安全 (3)
    ("TC-SEC-001", "/文件上传", "安全测试", 1, "上传_伪PHP木马", "testuser01", "file=1.jpg内容为<?php...?>", "code=40001"),
    ("TC-SEC-002", "/头像上传", "安全测试", 1, "头像_双扩展名绕过", "testuser01", "file=1.jpg.php", "code=40001"),
    ("TC-SEC-003", "/头像上传", "安全测试", 1, "头像_越权修改别人", "testuser02", "POST /file/upload/avatar userId=别人", "只改本人"),
]

# =============== 生成 4 份 xlsx ===============
def build_xlsx(name, cases, out_dir):
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append(HEADERS)
    for c_idx in range(1, len(HEADERS)+1):
        cell = ws.cell(row=1, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for i, c in enumerate(cases, start=1):
        ws.append([c[0], PRODUCT, c[1], c[2], c[3], c[4], c[5], c[6], c[7], "", "", name, ""])

    for r in range(2, ws.max_row+1):
        for c_idx in range(1, len(HEADERS)+1):
            cell = ws.cell(row=r, column=c_idx)
            cell.alignment = left
            cell.border = border
            if c_idx in (3, 4, 5):
                cell.alignment = center
            if c_idx == 6:
                cell.font = Font(bold=True)

    col_widths = [14, 18, 16, 14, 6, 30, 28, 38, 32, 14, 8, 10, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    for r in range(2, ws.max_row+1):
        ws.row_dimensions[r].height = 80
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

    # 统计 sheet
    ws2 = wb.create_sheet("用例统计")
    ws2.append(["测试类型", "用例数量"])
    type_count = {}
    for c in cases:
        type_count[c[2]] = type_count.get(c[2], 0) + 1
    for k, v in sorted(type_count.items(), key=lambda x: -x[1]):
        ws2.append([k, v])
    ws2.append(["合计", len(cases)])
    for r in range(1, ws2.max_row+1):
        for c_idx in range(1, 3):
            cell = ws2.cell(row=r, column=c_idx)
            cell.alignment = center
            cell.border = border
            if r == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 12

    out_path = out_dir / f"{name}_脚本与截图/软件测试测试用例.xlsx"
    wb.save(out_path)
    print(f"  {name}: {len(cases)} 条 → {out_path}")
    return len(cases)

# 执行
members = [
    ("朱远亮", zhuyuanliang_cases),
    ("李冠燃", liguanran_cases),
    ("李坤纬", likunwei_cases),
    ("林景彬", linjingbin_cases),
]

print("=== 生成 4 份分模块 xlsx ===")
total = 0
for name, cases in members:
    n = build_xlsx(name, cases, BASE)
    total += n
print(f"\n总计: {total} 条")

# =============== 生成汇总 xlsx ===============
print("\n=== 生成汇总 xlsx ===")
wb = Workbook()
ws = wb.active
ws.title = "测试用例汇总"
ws.append(HEADERS)
for c_idx in range(1, len(HEADERS)+1):
    cell = ws.cell(row=1, column=c_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

for name, cases in members:
    for c in cases:
        ws.append([c[0], PRODUCT, c[1], c[2], c[3], c[4], c[5], c[6], c[7], "", "", name, ""])

for r in range(2, ws.max_row+1):
    for c_idx in range(1, len(HEADERS)+1):
        cell = ws.cell(row=r, column=c_idx)
        cell.alignment = left
        cell.border = border
        if c_idx in (3, 4, 5):
            cell.alignment = center
        if c_idx == 6:
            cell.font = Font(bold=True)

col_widths = [14, 18, 16, 14, 6, 30, 28, 38, 32, 14, 8, 10, 12]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 30
for r in range(2, ws.max_row+1):
    ws.row_dimensions[r].height = 80
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

summary_path = BASE / "StarPicture_测试用例.xlsx"
wb.save(summary_path)
print(f"  汇总: {total} 条 → {summary_path}")

# =============== 更新评分表 ===============
print("\n=== 更新评分表 ===")
import shutil
scoring_src = BASE / "StarPicture_评分表.xlsx"
if scoring_src.exists():
    wb = load_workbook(scoring_src, data_only=False)
    ws = wb.active

    # 4 人用例数
    member_totals = {name: len(cases) for name, cases in members}

    for i, (name, cases) in enumerate(members):
        r = 3 + i
        if ws.cell(r, 1).value or ws.cell(r, 2).value:
            # 按类型统计
            type_count = {}
            for c in cases:
                type_count[c[2]] = type_count.get(c[2], 0) + 1

            # 列 9-15: 功能/性能/接口/安全/自动/单元/兼容
            ws.cell(r, 9).value = type_count.get('功能测试', 0)
            ws.cell(r, 10).value = type_count.get('性能测试', 0)
            ws.cell(r, 11).value = type_count.get('接口测试', 0)
            ws.cell(r, 12).value = type_count.get('安全测试', 0)
            ws.cell(r, 13).value = 0  # 自动
            ws.cell(r, 14).value = 0  # 单元
            ws.cell(r, 15).value = 0  # 兼容
            # 列 24 备注
            ws.cell(r, 24).value = f"{name} 实际 {len(cases)} 条"

    wb.save(scoring_src)
    print(f"  评分表已更新 → {scoring_src}")

# 验证
print("\n=== 验证 ===")
for name, cases in members:
    print(f"  {name}: {len(cases)} 条")
print(f"  总计: {total} 条")
print(f"\n完成！{total} 条用例，覆盖 58 个真实 API。")