"""
按新要求更新评分表和汇总 xlsx：
- 列 10 = 功能（10 条）
- 列 11 = 性能（1 条）
- 列 12 = 接口（1 条）
- 列 13 = 安全（1 条）
- 列 14-16 = 自动/单元/兼容（0 条）
- 总 13 条 × 4 = 52 条
"""
from openpyxl import load_workbook
from pathlib import Path

BASE = Path("D:/code/StarPicture/docs/test")
scoring = BASE / "StarPicture_评分表.xlsx"
summary = BASE / "StarPicture_测试用例.xlsx"

# ============ 1. 更新评分表 ============
wb = load_workbook(scoring)
ws = wb.active

NAME_MAP = {
    "韩东晓": "朱远亮",
    "陈雨腾": "李冠燃",
    "甘松明": "李坤纬",
    "吴泽波": "林景彬",
}

# 列 10-16: 功能 性能 接口 安全 自动 单元 兼容
NEW_CASES = {
    "朱远亮": (10, 1, 1, 1, 0, 0, 0),
    "李冠燃": (10, 1, 1, 1, 0, 0, 0),
    "李坤纬": (10, 1, 1, 1, 0, 0, 0),
    "林景彬": (10, 1, 1, 1, 0, 0, 0),
}

for r in range(3, ws.max_row + 1):
    leader = ws.cell(r, 1).value
    member = ws.cell(r, 2).value
    key = leader or member
    if key in NAME_MAP:
        new_name = NAME_MAP[key]
        # 改名字
        if leader == key:
            ws.cell(r, 1).value = new_name
            if member == key:
                ws.cell(r, 2).value = new_name
        else:
            ws.cell(r, 2).value = new_name
        # 改用例数（列 10-16）
        gong, xing, jie, an, zi, dan, rong = NEW_CASES[new_name]
        ws.cell(r, 10).value = gong
        ws.cell(r, 11).value = xing
        ws.cell(r, 12).value = jie
        ws.cell(r, 13).value = an
        ws.cell(r, 14).value = zi
        ws.cell(r, 15).value = dan
        ws.cell(r, 16).value = rong
        ws.cell(r, 24).value = f"{new_name} 实际 13 条（功能 10 + 性能 1 + 接口 1 + 安全 1）"

wb.save(scoring)
print(f"已更新评分表: {scoring}")

# ============ 2. 验证评分表 ============
print("\n=== 评分表验证 ===")
wb = load_workbook(scoring)
ws = wb.active
for r in range(3, 7):
    name = ws.cell(r,1).value or ws.cell(r,2).value or ''
    gx = ws.cell(r,3).value
    cells = [ws.cell(r,c).value for c in range(10,17)]
    note = ws.cell(r,24).value or ''
    print(f"  {name:<8} 贡献度={gx} | 功能{cells[0]} 性能{cells[1]} 接口{cells[2]} 安全{cells[3]} 自动{cells[4]} 单元{cells[5]} 兼容{cells[6]}")
    print(f"    备注: {note}")

# ============ 3. 验证汇总 xlsx ============
print("\n=== 汇总 xlsx 验证 ===")
wb = load_workbook(summary, data_only=True)
ws = wb["测试用例汇总"]
total = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[0])
print(f"汇总总数: {total}")
from collections import defaultdict
type_count = defaultdict(int)
member_count = defaultdict(int)
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        type_count[row[3]] += 1
        member_count[row[11]] += 1
print("按类型:")
for k, v in sorted(type_count.items(), key=lambda x: -x[1]):
    print(f"  {k}: {v}")
print("按人员:")
for k, v in sorted(member_count.items()):
    print(f"  {k}: {v}")