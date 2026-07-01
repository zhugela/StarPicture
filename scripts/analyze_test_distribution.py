"""
按贡献权重 1.1:1:1:1 分配全部 104 条测试用例给 4 人
- 朱远亮(1.1): 28 条
- 李冠燃(1.0): 25 条
- 李坤纬(1.0): 25 条
- 林景彬(1.0): 26 条

分配规则：
1. 每人保留自己模块的功能测试（11条）
2. 性能/接口/安全 3 种类型，按权重从其他模块匀给 朱远亮 多 3 条
3. 最终每人都有功能+性能+接口+安全 4 种类型
"""

from openpyxl import load_workbook

# 加载汇总 xlsx（104 条）
wb = load_workbook('D:/code/StarPicture/docs/内娱图库_海蒂与爷爷_朱远亮_18144610287/StarPicture_测试用例.xlsx')
ws = wb['测试用例汇总']

# 读取所有用例
all_cases = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        all_cases.append({
            'name': row[0],
            'module': row[1],
            'type': row[2],
            'id': row[3],
            'desc': row[4],
            'priority': row[5],
        })

print(f"总用例数: {len(all_cases)}")

# 按类型分组
by_type = {}
for c in all_cases:
    t = c['type']
    if t not in by_type:
        by_type[t] = []
    by_type[t].append(c)

print("按类型分布:")
for t, cases in sorted(by_type.items()):
    print(f"  {t}: {len(cases)} 条")

# 每种类型的用例数
func_cases = by_type.get('功能测试', [])
perf_cases = by_type.get('性能测试', [])
api_cases = by_type.get('接口测试', [])
sec_cases = by_type.get('安全测试', [])

print(f"\n功能: {len(func_cases)}, 性能: {len(perf_cases)}, 接口: {len(api_cases)}, 安全: {len(sec_cases)}")
print(f"总计: {len(func_cases) + len(perf_cases) + len(api_cases) + len(sec_cases)}")
