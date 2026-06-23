"""
替换林景彬的"微信公众号门户"为"头像上传"
要更新的文件：
1. 林景彬_脚本与截图/软件测试测试用例.xlsx
2. 林景彬_脚本与截图/README.md
3. 林景彬_脚本与截图/TUTORIAL_保姆级教程.md
4. 林景彬_脚本与截图/功能测试/file_wxmp_functional.postman_collection.json
5. 林景彬_脚本与截图/接口测试/file_wxmp_api.postman_collection.json
6. 林景彬_脚本与截图/性能测试/file_upload_50concurrent_1MB.jmx（保留）
7. 林景彬_脚本与截图/安全测试/文件_公众号模块-安全报告模板.docx
8. 顶层 StarPicture_测试用例.xlsx（汇总）
"""
import json
import shutil
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from docx import Document

BASE = Path("D:/code/StarPicture/docs/test")
LJB = BASE / "林景彬_脚本与截图"

# 新模块：头像上传 + 获取当前用户（也属于 user/avatar 模块）
# 用例编号：TC-AT-xxx (avatar) + TC-GC-xxx (get current)

PRODUCT = "内娱图库StarPicture"

# ============ 1. 重新生成 林景彬的用例 xlsx ============
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="305496")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(border_style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["用例编号", "所属产品", "所属模块", "用例类型", "优先级", "用例标题",
           "前置条件", "步骤", "预期结果", "实测结果", "结论", "测试人员", "测试时间"]

ljb_cases = [
    # 功能：头像上传
    ("TC-AT-001", "/头像上传", "功能测试", 1, "头像上传_jpg_512KB", "testuser01 已登录",
     "1. POST /file/upload/avatar file=512KB_jpg（Content-Type: multipart/form-data）",
     "返回code=0, userAvatar 字段已更新为新 URL"),
    ("TC-AT-002", "/头像上传", "功能测试", 1, "头像上传_png", "testuser01 已登录",
     "1. POST /file/upload/avatar file=512KB_png",
     "返回code=0, picFormat=png"),
    ("TC-AT-003", "/头像上传", "功能测试", 1, "头像上传_超过2MB_应失败", "testuser01 已登录",
     "1. POST /file/upload/avatar file=5MB_jpg",
     "返回code=40001 提示文件过大"),
    ("TC-AT-004", "/头像上传", "功能测试", 2, "头像上传_非图片_应失败", "testuser01 已登录",
     "1. POST /file/upload/avatar file=test.pdf",
     "返回code=40001 提示格式错误"),
    ("TC-AT-005", "/头像上传", "功能测试", 2, "头像上传_空文件_应失败", "testuser01 已登录",
     "1. POST /file/upload/avatar file=0byte",
     "返回code=40001 提示文件为空"),
    # 功能：获取当前用户
    ("TC-GC-001", "/获取当前用户", "功能测试", 1, "获取当前用户_已登录", "testuser01 已登录",
     "1. 携带 Cookie\n2. GET /user/get/login",
     "返回 testuser01 的 userVo，含 userAvatar"),
    ("TC-GC-002", "/获取当前用户", "功能测试", 1, "获取当前用户_未登录", "无 Cookie",
     "1. 不带 Cookie\n2. GET /user/get/login",
     "返回code=40100 提示未登录"),
    ("TC-GC-003", "/获取当前用户", "功能测试", 2, "获取当前用户_Cookie过期", "Cookie 过期",
     "1. 设置 Cookie 过期时间为过去\n2. GET /user/get/login",
     "返回code=40100"),
    ("TC-GC-004", "/获取当前用户", "功能测试", 2, "获取当前用户_Cookie伪造", "伪造 Cookie",
     "1. 手动设置 userId=999999 的 Cookie\n2. GET /user/get/login",
     "返回code=40100 或 code=40001，不应返回该用户"),
    ("TC-GC-005", "/获取当前用户", "功能测试", 2, "获取当前用户_头像URL正确", "testuser01 已上传头像",
     "1. 先上传头像\n2. GET /user/get/login",
     "返回的 userVo 里 userAvatar 是新上传的 URL"),
    # 性能
    ("TC-PERF-001", "/头像上传", "性能测试", 2, "头像上传_50并发", "已安装 JMeter",
     "1. JMeter 50 线程 5s 内启动\n2. POST /file/upload/avatar file=512KB_jpg",
     "P95 < 2s, 错误率 < 5%"),
    # 接口
    ("TC-API-001", "/头像上传", "接口测试", 2, "头像上传_缺multipart边界", "testuser01 已登录",
     "1. 发送不完整 multipart 请求\n2. POST /file/upload/avatar",
     "返回 400 或 415"),
    # 安全
    ("TC-SEC-001", "/头像上传", "安全测试", 1, "头像上传_伪PHP木马", "testuser01 已登录",
     "1. file=1.jpg 内容为 <?php system($_GET['c']);?>",
     "返回code=40001 拒绝上传"),
]

wb = Workbook()
ws = wb.active
ws.title = "测试用例"
ws.append(HEADERS)
for c_idx in range(1, len(HEADERS)+1):
    cell = ws.cell(row=1, column=c_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

for c in ljb_cases:
    row = [c[0], PRODUCT, c[1], c[2], c[3], c[4], c[5], c[6], c[7], "", "", "林景彬", ""]
    ws.append(row)

for r in range(2, ws.max_row+1):
    for c_idx in range(1, len(HEADERS)+1):
        cell = ws.cell(row=r, column=c_idx)
        cell.alignment = left
        cell.border = border
        if c_idx in (3, 4, 5):
            cell.alignment = center
        if c_idx == 6:
            cell.font = Font(bold=True)

for i, w in enumerate([14, 18, 16, 14, 6, 30, 28, 38, 32, 14, 8, 10, 12], start=1):
    ws.column_dimensions[chr(64+i)].width = w
ws.row_dimensions[1].height = 30
for r in range(2, ws.max_row+1):
    ws.row_dimensions[r].height = 80
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:M{ws.max_row}"

# 统计 sheet
ws2 = wb.create_sheet("用例统计")
ws2.append(["项目", "数据"])
ws2.append(["负责模块", "头像上传 + 获取当前用户"])
ws2.append(["用例总数", len(ljb_cases)])
ws2.append([])
ws2.append(["测试类型", "用例数"])
type_count = {}
for c in ljb_cases:
    type_count[c[2]] = type_count.get(c[2], 0) + 1
for k, v in sorted(type_count.items(), key=lambda x: -x[1]):
    ws2.append([k, v])
for r in range(1, ws2.max_row+1):
    for c_idx in range(1, 3):
        cell = ws2.cell(row=r, column=c_idx)
        cell.alignment = center
        cell.border = border
        if r == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 12

out = LJB / "软件测试测试用例.xlsx"
wb.save(out)
print(f"1. 已生成: {out} (13 条用例)")

# ============ 2. 重新生成 功能测试 Postman ============
def make_col(name, items):
    return {
        "info": {
            "name": name,
            "description": "导入：Postman → File → Import → Upload Files → 选此 json\n跑法：左侧点集合 → 选每个用例 → 右侧 Send → 截图为证",
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"
        },
        "item": items,
        "variable": [
            {"key": "baseUrl", "value": "http://localhost:8123/api"},
            {"key": "token", "value": ""},
        ]
    }

def req(name, method, path, body=None, desc=""):
    item = {
        "name": name,
        "request": {
            "method": method,
            "header": [{"key": "Content-Type", "value": "application/json"}],
            "url": {
                "raw": "{{baseUrl}}" + path,
                "host": ["{{baseUrl}}"],
                "path": [p for p in path.split("/") if p]
            },
            "description": desc
        },
        "response": []
    }
    if body is not None:
        item["request"]["body"] = {"mode": "raw", "raw": json.dumps(body, ensure_ascii=False, indent=2)}
    return item

functional_items = [
    req("TC-AT-001_头像上传_jpg_512KB", "POST", "/file/upload/avatar", None,
        "form-data 模式：file=512KB_jpg（要先登录拿 Cookie）"),
    req("TC-AT-002_头像上传_png", "POST", "/file/upload/avatar", None,
        "form-data 模式：file=512KB_png"),
    req("TC-AT-003_头像上传_超过2MB", "POST", "/file/upload/avatar", None,
        "form-data 模式：file=5MB_jpg"),
    req("TC-AT-004_头像上传_非图片", "POST", "/file/upload/avatar", None,
        "form-data 模式：file=test.pdf"),
    req("TC-AT-005_头像上传_空文件", "POST", "/file/upload/avatar", None,
        "form-data 模式：file=0byte.jpg"),
    req("TC-GC-001_获取当前用户_已登录", "GET", "/user/get/login"),
    req("TC-GC-002_获取当前用户_未登录", "GET", "/user/get/login"),
    req("TC-GC-003_获取当前用户_Cookie过期", "GET", "/user/get/login"),
    req("TC-GC-004_获取当前用户_Cookie伪造", "GET", "/user/get/login"),
    req("TC-GC-005_获取当前用户_头像URL正确", "GET", "/user/get/login"),
]
functional_col = make_col("StarPicture_头像_获取用户模块_功能测试", functional_items)
out = LJB / "功能测试/file_wxmp_functional.postman_collection.json"
out.write_text(json.dumps(functional_col, ensure_ascii=False, indent=2), encoding='utf-8')
print(f"2. 已生成: {out}")

# ============ 3. 重新生成 接口测试 Postman ============
api_items = [
    req("API-001_头像上传_缺multipart边界", "POST", "/file/upload/avatar", None,
        "Headers 改成 application/xml + Body raw 写乱码"),
    req("API-002_头像上传_Content-Type错误", "POST", "/file/upload/avatar", None,
        "Content-Type=application/json + Body 二进制"),
    req("API-003_头像_无文件_应失败", "POST", "/file/upload/avatar", None,
        "不传 file 字段"),
    req("API-004_获取用户_无Cookie", "GET", "/user/get/login"),
    req("API-005_获取用户_方法错误", "POST", "/user/get/login"),
]
api_col = make_col("StarPicture_头像_获取用户模块_接口测试", api_items)
out = LJB / "接口测试/file_wxmp_api.postman_collection.json"
out.write_text(json.dumps(api_col, ensure_ascii=False, indent=2), encoding='utf-8')
print(f"3. 已生成: {out}")

# ============ 4. 重新生成 README.md ============
readme = """# 林景彬_脚本与截图

> **姓名**：林景彬
> **学号**：2310820051
> **负责模块**：file + user（头像上传 + 获取当前用户）
> **用例总数**：13 条（2 个功能点 × 5 条 + 性能 1 + 接口 1 + 安全 1）
> **用例文件**：[软件测试测试用例.xlsx](./软件测试测试用例.xlsx)

## 我的工作清单

| 类型 | 用例数 | 工作 |
|---|---|---|
| 功能测试 | 10 | 2 个功能点 |
| 性能测试 | 1 | 头像上传 50 并发（JMeter） |
| 接口测试 | 1 | 头像上传缺 multipart 边界（Postman） |
| 安全测试 | 1 | 伪 PHP 木马上传（BurpSuite） |

## 我的 2 个功能点

1. **头像上传**（TC-AT-001 ~ TC-AT-005）
   - jpg/png/超2MB/非图片/空文件
2. **获取当前用户**（TC-GC-001 ~ TC-GC-005）
   - 已登录/未登录/Cookie过期/Cookie伪造/头像URL正确

## 5 个子目录具体放什么

```
林景彬_脚本与截图/
├── 软件测试测试用例.xlsx              ← 13 条用例清单
├── 功能测试/                          ← 10 张 Postman 用例截图
│   └── file_wxmp_functional.postman_collection.json
├── 安全测试/                          ← PDF + .scan
│   ├── 文件_公众号模块-安全报告模板.docx
│   ├── 文件_公众号模块-安全报告.pdf                      ← 待办：填完模板导出 PDF
│   └── 文件_公众号模块.scan                    ← 待办：BurpSuite 扫出
├── 性能测试/                          ← .jmx + 截图
│   └── file_upload_50concurrent_1MB.jmx
├── 接口测试/                          ← .json + 截图
│   └── file_wxmp_api.postman_collection.json
├── TUTORIAL_保姆级教程.md             ← 详细操作步骤
└── README.md                          ← 本文件
```

## 具体待办（按时间）

### Day 1（6/17 下午 ~ 6/18 中午）
- [ ] 评审自己写的 13 条用例
- [ ] 装 Postman + JMeter + BurpSuite

### Day 2（6/18 上午 ~ 6/19 中午）
- [ ] 跑 10 条功能用例，每条截 1 张图，存到 `功能测试/`
- [ ] 跑 1 条性能用例（JMeter），截图 Summary Report
- [ ] 跑 1 条接口用例（Postman），截图
- [ ] 跑 1 条安全用例（BurpSuite），截图

### Day 3（6/19 上午 ~ 6/19 晚）
- [ ] WPS 打开 `文件_公众号模块-安全报告模板.docx`，填实际结果 → 导出 PDF
- [ ] BurpSuite 扫后端，导出 `.scan`
- [ ] 回归测试
- [ ] 把缺陷填到 `软件测试报告.md`

## 不会做的看 TUTORIAL_保姆级教程.md

---

**有问题群里 @ 林景彬**
"""
(LJB / "README.md").write_text(readme, encoding='utf-8')
print(f"4. 已生成: README.md")

# ============ 5. 重新生成 TUTORIAL 保姆级教程 ============
tutorial = """# 林景彬 · 保姆级教程（零基础也能跟做）

> **你的名字**：林景彬
> **学号**：2310820051
> **负责模块**：file + user（头像上传 + 获取当前用户）
> **用例总数**：13 条（10 功能 + 1 性能 + 1 接口 + 1 安全）
> **目标**：跟着做，3 天内交齐所有文件
> **难度**：⭐⭐⭐

---

## 📋 0. 你要交的所有东西

打开你的文件夹 `D:\\code\\StarPicture\\docs\\test\\林景彬_脚本与截图\\`，里面已经有：

```
林景彬_脚本与截图/
├── README.md                            ← 说明
├── TUTORIAL_保姆级教程.md               ← 你正在看的教程
├── 软件测试测试用例.xlsx                 ← 13 条用例（不要改）
│
├── 功能测试/                            ← 你要做：跑 10 个用例 + 截 10 张图
│   └── file_wxmp_functional.postman_collection.json
│
├── 性能测试/                            ← 你要做：跑 JMeter + 截图
│   └── file_upload_50concurrent_1MB.jmx
│
├── 接口测试/                            ← 你要做：跑 1 个用例 + 截图
│   └── file_wxmp_api.postman_collection.json
│
└── 安全测试/                            ← 你要做：填报告 + 写 .scan
    └── 文件_公众号模块-安全报告模板.docx
```

**你最后要交的是**：

| 目录 | 文件 | 数量 |
|---|---|---|
| 功能测试/ | TC-AT-001_xxx.png ~ TC-GC-005_xxx.png | **10 张** |
| 性能测试/ | summary_report_xxx.png | **1 张** |
| 接口测试/ | API-001_xxx.png | **1 张** |
| 安全测试/ | 文件_公众号模块-安全报告.pdf | **1 个** |
| 安全测试/ | 文件_公众号模块.scan | **1 个** |

---

## ⏱️ 1. Day 1 下午：装软件（1 小时）

### 1.1 装 Postman

**步骤**：

1. 打开浏览器，访问 https://www.postman.com/downloads/
2. 点 "Windows 64-bit" 下载（约 200MB）
3. 双击 `Postman-win64-Setup.exe` 安装
4. 安装完自动打开 Postman
5. 第一次打开注册：点 "Create Account" → 用邮箱注册 → 登录
6. **看到主界面就 OK**

**验证**：打开 Postman，左上角应该有 "Workspaces"，右侧应该有 "Send" 按钮。

### 1.2 装 JMeter

**步骤**：

1. 访问 https://jmeter.apache.org/download_jmeter.cgi
2. 下载 `apache-jmeter-5.6.3.zip`（约 80MB）
3. **右键 → 解压到 `D:\\Program Files\\`**
4. 解压完会有 `D:\\Program Files\\apache-jmeter-5.6.3\\bin\\jmeter.bat`

**配环境变量**：

1. `Win + R` → `sysdm.cpl` → 回车
2. 切到 "高级" 标签页
3. 点底部的 "环境变量(N)..."
4. "系统变量" → "新建(W)..."：
   - 变量名：`JMETER_HOME`
   - 变量值：`D:\\Program Files\\apache-jmeter-5.6.3`
   - 点确定
5. 找 `Path` → "编辑(I)..." → "新建(N)" → 输入 `%JMETER_HOME%\\bin` → 确定
6. 一路确定关掉所有窗口

**验证**：

1. `Win + R` → `cmd` → 回车
2. 在黑窗口输入 `jmeter`，回车
3. **期待弹出 JMeter 图形窗口**

### 1.3 装 BurpSuite

**步骤**：

1. 访问 https://portswigger.net/burp/communitydownload
2. 选 "Windows 64-bit" → 下载（约 500MB）
3. 双击 `burp-suite-windows-installer.exe` 安装
4. 双击桌面 "Burp Suite Community Edition" 打开
5. 弹窗 "Welcome" → Next → 接受条款 → 选 "Temporary project" → Start Burp

### 1.4 验证后端能起来

群里问林景彬：**"后端起来了没？"**。如果他说"起了"，打开浏览器访问 `http://localhost:8123/api/doc.html`，期待看到 Knife4j 接口文档。

---

## ⏱️ 2. Day 2 上午：跑功能测试（1 小时）

### 2.1 打开 Postman + 导入集合

**步骤**：

1. 打开 Postman
2. 顶部菜单 **File → Import**
3. 点 **Upload Files**
4. 选文件：`D:\\code\\StarPicture\\docs\\test\\林景彬_脚本与截图\\功能测试\\file_wxmp_functional.postman_collection.json`
5. 点右下角 **Import**
6. 左侧出现 `StarPicture_头像_获取用户模块_功能测试` 集合
7. 点开它，看到 10 个用例

### 2.2 配置环境变量

**步骤**：

1. Postman 右上角齿轮图标 → **+** 创建
2. **Environment Name**：`StarPicture本地`
3. 表格：
   - Variable：`baseUrl`
   - Type：`default`
   - Initial Value：`http://localhost:8123/api`
4. **Save**
5. **右上角下拉框** 选 `StarPicture本地`

### 2.3 ⚠️ 先登录拿 Cookie（重要！）

**步骤**：

1. 左侧创建一个**临时请求**：
   - 方法：`POST`
   - URL：`{{baseUrl}}/user/login`
   - Body → raw → JSON：
     ```json
     {"userAccount": "testuser01", "userPassword": "12345678"}
     ```
2. 点 **Send**
3. 看右下 Response：`{"code": 0, ...}` 表示登录成功
4. **Postman 自动保存了 Cookie**，后续请求会带上

### 2.4 跑第 1 个用例：TC-AT-001 头像上传 jpg 512KB

**步骤**：

1. 左侧集合展开 → 找到 **TC-AT-001_头像上传_jpg_512KB** → 点开
2. 右侧中部显示：`POST  http://localhost:8123/api/file/upload/avatar`
3. 中间偏下选 **Body** 标签 → 选 **form-data** → key 填 `file`（File 类型）→ 选一张 512KB 的 jpg
4. 右侧 **Send** 按钮（蓝色）→ 点一下
5. 期待右下角 Response：
   - Status: `200 OK`
   - Body: `{"code": 0, "data": {"userAvatar": "http://..."}}`

### 2.5 截图 ⭐

**步骤**：

1. **不要关闭 Postman**（保持窗口原样）
2. 按 **`Win + Shift + S`**
3. 选 **"窗口截图"**
4. 鼠标移到 Postman → 自动框选
5. 点一下，截图保存到剪贴板
6. 打开 **画图**，按 `Ctrl + V` 粘贴
7. 另存为 PNG 到：`D:\\code\\StarPicture\\docs\\test\\林景彬_脚本与截图\\功能测试\\TC-AT-001_头像上传_jpg_512KB.png`

### 2.6 跑剩下 9 个用例

| 用例 | 期望结果 |
|---|---|
| TC-AT-002 头像上传 png | code=0, picFormat=png |
| TC-AT-003 头像上传 超过2MB | code=40001 |
| TC-AT-004 头像上传 非图片 | code=40001 |
| TC-AT-005 头像上传 空文件 | code=40001 |
| TC-GC-001 获取当前用户_已登录 | code=0, 返回 userVo |
| TC-GC-002 获取当前用户_未登录 | code=40100 |
| TC-GC-003 获取当前用户_Cookie过期 | code=40100 |
| TC-GC-004 获取当前用户_Cookie伪造 | code=40100/40001 |
| TC-GC-005 获取当前用户_头像URL正确 | code=0, userAvatar 是新 URL |

**每个用例截图 1 张，命名 `TC-XXX-NNN_标题.png`，存到 `功能测试/` 目录。** 10 张图都要截。

---

## ⏱️ 3. Day 2 下午：跑性能测试（30 分钟）

### 3.1 打开 JMeter + 加载 .jmx

**步骤**：

1. `Win + R` → `cmd` → 输入 `jmeter` → 回车
2. JMeter 主窗口弹出
3. 顶部菜单 **File → Open**
4. 选文件：`D:\\code\\StarPicture\\docs\\test\\林景彬_脚本与截图\\性能测试\\file_upload_50concurrent_1MB.jmx`
5. 左侧看到：测试计划 → 线程组 → HTTP 请求

### 3.2 修改线程数和目标

**步骤**：

1. 左侧点 **线程组** → 右侧：
   - `Number of Threads (users)`：`50`
   - `Ramp-up period`：填 `5`
   - `Loop Count`：填 `1`
2. 左侧点 **HTTP 请求** → 右侧：
   - `Server Name or IP`：`localhost`
   - `Port Number`：`8123`
   - `Method`：`POST`
   - `Path`：`/file/upload/avatar`

### 3.3 添加 Summary Report + 运行

**步骤**：

1. 左侧点 **HTTP 请求** → **Edit → Add → Listener → Summary Report**
2. 顶部 **Run → Start**（或绿色 ▶）
3. 观察右下角显示 "50/50"
4. 等 10-30 秒，跑完显示 "0/50"
5. 看 Summary Report：
   - `# Samples`：50
   - `Average`：应 < 2000ms
   - `Error %`：应 < 5%

### 3.4 截图 ⭐

按 `Win+Shift+S` → 框选 JMeter → 保存到 `性能测试/summary_report_头像上传50并发.png`

---

## ⏱️ 4. Day 2 下午：跑接口测试（15 分钟）

### 4.1 导入接口集合

Postman File → Import → Upload Files → 选 `接口测试\\file_wxmp_api.postman_collection.json`

### 4.2 跑 API-001 头像上传缺 multipart 边界

**步骤**：

1. 左侧点 **API-001_头像上传_缺multipart边界**
2. 右侧选 **Headers** → 把 Content-Type 改成 `application/xml`（错的）
3. 切到 **Body** → raw → XML → 写乱码
4. 点 **Send**
5. 期待 Response：`415` 或 `400`

### 4.3 截图

保存到 `接口测试/API-001_头像上传_缺multipart边界.png`

---

## ⏱️ 5. Day 3 上午：跑安全测试（2 小时）

### 5.1 写 PDF

**步骤**：

1. 打开 WPS 文字
2. **文件 → 打开** → 选 `安全测试\\文件_公众号模块-安全报告模板.docx`
3. 找所有 `[待填]`，替换为实际结果
4. **文件 → 导出为 PDF** → 文件名 `文件_公众号模块-安全报告.pdf` → 保存到 `安全测试/`

### 5.2 生成 .scan（用 BurpSuite）

**步骤**：

1. 双击 Burp Suite Community Edition 打开
2. 左侧 **Proxy** → 顶部 **Intercept is on** → **点一下变 off**
3. Chrome 装 FoxyProxy 扩展（已发过教程）
4. 浏览器访问 `http://localhost:8123/api/file/upload/avatar`
5. 上传 `1.jpg`（内容是 `<?php system($_GET['c']);?>`，记事本写保存）
6. 后端应返回 `code=40001`
7. BurpSuite **Target** 标签 → 找到 `/api/file/upload/avatar` → 右键 **Save item** → 选 `.scan` 格式 → 保存

### 5.3 装不上 BurpSuite？

写 Markdown 改名为 `文件_公众号模块.scan.md`，记录手动测试结果。

---

## ⏱️ 6. Day 3 下午：最后检查（30 分钟）

打开你的文件夹，确认：

```
林景彬_脚本与截图/
├── 功能测试/                        ← 应该有 10 张 png
├── 性能测试/                        ← 应该有 1 张 png
├── 接口测试/                        ← 应该有 1 张 png
├── 安全测试/                        ← 应该有 PDF + .scan
```

**12 张截图清单**：

| 文件名 |
|---|
| 功能测试/TC-AT-001_头像上传_jpg_512KB.png |
| 功能测试/TC-AT-002_头像上传_png.png |
| 功能测试/TC-AT-003_头像上传_超过2MB.png |
| 功能测试/TC-AT-004_头像上传_非图片.png |
| 功能测试/TC-AT-005_头像上传_空文件.png |
| 功能测试/TC-GC-001_获取当前用户_已登录.png |
| 功能测试/TC-GC-002_获取当前用户_未登录.png |
| 功能测试/TC-GC-003_获取当前用户_Cookie过期.png |
| 功能测试/TC-GC-004_获取当前用户_Cookie伪造.png |
| 功能测试/TC-GC-005_获取当前用户_头像URL正确.png |
| 性能测试/summary_report_头像上传50并发.png |
| 接口测试/API-001_头像上传_缺multipart边界.png |

**群里汇报**：准备好后告诉群里"我的搞定了"。

---

## 🆘 出问题

| 问题 | 找谁 |
|---|---|
| 后端起不来 | 林景彬 |
| 不知道某步怎么点 | 群里截图问 |
| BurpSuite 装不上 | 写 .scan.md |

---

**教程作者**：朱远亮（组长）
**最后更新**：2026-06-19
"""
(LJB / "TUTORIAL_保姆级教程.md").write_text(tutorial, encoding='utf-8')
print(f"5. 已生成: TUTORIAL_保姆级教程.md")

# ============ 6. 更新顶层汇总 xlsx ============
summary = BASE / "StarPicture_测试用例.xlsx"
wb = load_workbook(summary)
ws = wb["测试用例汇总"]

# 删林景彬的旧数据
rows_to_delete = []
for r in range(2, ws.max_row+1):
    if ws.cell(r, 11).value == "林景彬":
        rows_to_delete.append(r)
for r in reversed(rows_to_delete):
    ws.delete_rows(r)

# 重新添加林景彬的新用例
for c in ljb_cases:
    row = [c[0], PRODUCT, c[1], c[2], c[3], c[4], c[5], c[6], c[7], "", "", "林景彬", ""]
    ws.append(row)

wb.save(summary)
print(f"6. 已更新: 顶层汇总 xlsx（林景彬换成头像模块）")

print("\n全部完成！")