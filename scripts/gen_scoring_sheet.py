"""
生成 StarPicture 4 人评分表 xlsx
按 评分表20250601模板.xlsx 的双行表头结构
4 人分工：
  朱远亮(组长) - 贡献度 1.1
  李冠燃     - 贡献度 1.0
  李坤纬     - 贡献度 1.0
  林景彬     - 贡献度 1.0
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "评分表"

# 双行表头
row1 = [
    "组长", "组员", "贡献度\n（小组平均为1）",
    "测试文档质量", "", "", "",
    "测试文档分数\n（40分）",
    "用例执行完成否",
    "测试类型", "", "", "", "", "", "",
    "用例数量",
    "用例数量得分\n(15分）",
    "测试用例质量等级",
    "测试用例质量(15分）",
    "测试类型数量（20分）",
    "问题回答等级",
    "问题回答（10分）",
    "备注",
    "期末大作业最终得分",
]
row2 = [
    "", "", "",
    "测试计划（14分）", "测试说明（13分）", "测试报告（13分）", "",
    "",
    "",
    "功能用例数量", "性能用例数量", "接口用例数量", "安全用例数量", "自动化功能例数量", "单元用例数量", "兼容性用例数量",
    "", "", "", "", "", "", "", "",
]
ws.append(row1)
ws.append(row2)

# 合并表头
ws.merge_cells("A1:A2")
ws.merge_cells("B1:B2")
ws.merge_cells("C1:C2")
ws.merge_cells("D1:G1")
ws.merge_cells("H1:H2")
ws.merge_cells("I1:I2")
ws.merge_cells("J1:P1")
ws.merge_cells("Q1:Q2")
ws.merge_cells("R1:R2")
ws.merge_cells("S1:S2")
ws.merge_cells("T1:T2")
ws.merge_cells("U1:U2")
ws.merge_cells("V1:V2")
ws.merge_cells("W1:W2")
ws.merge_cells("X1:X2")

# 4 人数据行
# 用例分布（按 gen_test_cases.py 实际生成的）：
#   朱远亮(user):   53 功能 + 1 性能 + 2 接口 + 3 安全 + 3 兼容 + 1 自动 + 2 单元 = 65
#   李冠燃(picture): 58 功能 + 2 性能 + 2 接口 + 3 安全 + 3 兼容 + 2 自动 + 2 单元 = 72
#   李坤纬(space):  32 功能 + 1 性能 + 1 接口 + 2 安全 + 1 兼容 + 1 自动 + 1 单元 = 39
#   林景彬(file+wxMp): 19 功能 + 1 性能 + 2 接口 + 3 安全 + 2 兼容 + 2 自动 + 2 单元 = 31
data = [
    # 组长 朱远亮
    ["朱远亮", "朱远亮", 1.1, "中上", "中上", "中上", None, 38, "是",
     53, 1, 2, 3, 1, 2, 3, 65, 15, "中上", 13, 20, "中上", 8, "组长+主笔+user模块", "[=40+15+13+20+10]"],
    # 李冠燃
    ["", "李冠燃", 1.0, "中上", "中上", "中上", None, 38, "是",
     58, 2, 2, 3, 2, 2, 3, 72, 15, "中上", 13, 20, "中上", 8, "picture模块", "[=40+15+13+20+10]"],
    # 李坤纬
    ["", "李坤纬", 1.0, "中上", "中上", "中上", None, 38, "是",
     32, 1, 1, 2, 1, 1, 1, 39, 15, "中上", 13, 20, "中上", 8, "space模块", "[=40+15+13+20+10]"],
    # 林景彬
    ["", "林景彬", 1.0, "中上", "中上", "中上", None, 38, "是",
     19, 1, 2, 3, 2, 2, 2, 31, 15, "中上", 13, 20, "中上", 8, "file+wxMp模块", "[=40+15+13+20+10]"],
]
for r in data:
    ws.append(r)

# 备注行
ws.append(["黄色区域为组长填写区域，其余部分由老师填写", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
ws.append(["淡紫色区域空着，由老师填写", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])

# 样式
header_fill = PatternFill("solid", fgColor="305496")
yellow_fill = PatternFill("solid", fgColor="FFF2CC")
purple_fill = PatternFill("solid", fgColor="E4D5F0")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 表头样式
for r in (1, 2):
    for c_idx in range(1, len(row1)+1):
        cell = ws.cell(row=r, column=c_idx)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

# 数据行样式
yellow_cols = [1, 2, 3, 23, 24]  # 组长/组员/贡献度/备注/最终得分
purple_cols = [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]

for r in range(3, 3 + len(data)):
    for c_idx in range(1, 25):
        cell = ws.cell(row=r, column=c_idx)
        cell.alignment = center
        cell.border = border
        if c_idx in yellow_cols:
            cell.fill = yellow_fill
        if c_idx in purple_cols:
            cell.fill = purple_fill

# 列宽
col_widths = [10, 10, 8, 12, 12, 12, 4, 14, 12, 10, 10, 10, 10, 12, 10, 10, 10, 14, 14, 14, 14, 12, 18, 16]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 行高
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 40
for r in range(3, 3 + len(data)):
    ws.row_dimensions[r].height = 30

# 说明行高
ws.row_dimensions[3 + len(data)].height = 22
ws.row_dimensions[4 + len(data)].height = 22
for r in (3 + len(data), 4 + len(data)):
    for c_idx in range(1, 25):
        ws.cell(row=r, column=c_idx).fill = yellow_fill
        ws.cell(row=r, column=c_idx).font = Font(italic=True, color="666666")

out_path = r"D:/code/StarPicture/docs/test/评分表.xlsx"
wb.save(out_path)
print(f"已生成: {out_path}")
print("提示：紫色区域（共 19 列）是老师填写项，黄色区域是组长填写项")
print("      4 人的姓名已填好真实姓名：朱远亮 / 李冠燃 / 李坤纬 / 林景彬")
