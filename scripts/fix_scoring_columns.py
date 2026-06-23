"""
修复评分表列错位：
- 我之前写的列 10-16 → 实际应该是列 9-15
- 把功能/性能/接口/安全/自动/单元/兼容 7 列填到正确位置
- 列 16 (用例数量) 应该是 SUM 公式 = SUM(列9:列15)
"""
from openpyxl import load_workbook
from pathlib import Path

BASE = Path("D:/code/StarPicture/docs/test")
scoring = BASE / "StarPicture_评分表.xlsx"

wb = load_workbook(scoring)
ws = wb.active

NAME_MAP = {
    "韩东晓": "朱远亮",
    "陈雨腾": "李冠燃",
    "甘松明": "李坤纬",
    "吴泽波": "林景彬",
}

# 正确的列：9=功能 10=性能 11=接口 12=安全 13=自动 14=单元 15=兼容
# 每人：10 功能 + 1 性能 + 1 接口 + 1 安全 + 0 自动 + 0 单元 + 0 兼容
NEW_CASES = {
    "朱远亮": (10, 1, 1, 1, 0, 0, 0),
    "李冠燃": (10, 1, 1, 1, 0, 0, 0),
    "李坤纬": (10, 1, 1, 1, 0, 0, 0),
    "林景彬": (10, 1, 1, 1, 0, 0, 0),
}

for r in range(3, ws.max_row + 1):
    leader = ws.cell(r, 1).value
    member = ws.cell(r, 2).value
    key = leader or member
    if key in NAME_MAP:
        new_name = NAME_MAP[key]
        if leader == key:
            ws.cell(r, 1).value = new_name
            if member == key:
                ws.cell(r, 2).value = new_name
        else:
            ws.cell(r, 2).value = new_name
        # 改正确列：9-15
        gong, xing, jie, an, zi, dan, rong = NEW_CASES[new_name]
        ws.cell(r, 9).value = gong    # 功能
        ws.cell(r, 10).value = xing   # 性能
        ws.cell(r, 11).value = jie    # 接口
        ws.cell(r, 12).value = an     # 安全
        ws.cell(r, 13).value = zi     # 自动化
        ws.cell(r, 14).value = dan    # 单元
        ws.cell(r, 15).value = rong   # 兼容
        # 用例数量合计（列 16）= SUM(列9:列15)
        ws.cell(r, 16).value = f"=SUM(I{r}:O{r})"
        # 备注
        ws.cell(r, 24).value = f"{new_name} 实际 13 条（功能 10 + 性能 1 + 接口 1 + 安全 1）"

wb.save(scoring)
print(f"已修复评分表列错位")

# 验证
print("\n=== 验证（按列9-15正确读取） ===")
wb = load_workbook(scoring)
ws = wb.active
headers = ["功能", "性能", "接口", "安全", "自动", "单元", "兼容", "合计"]
for r in range(3, 7):
    name = ws.cell(r,1).value or ws.cell(r,2).value or ''
    cells = [ws.cell(r,c).value for c in range(9, 16)]
    total = ws.cell(r, 16).value
    note = ws.cell(r, 24).value or ''
    print(f"  {name:<8}", end=' ')
    for i, v in enumerate(cells):
        print(f"{headers[i]}={v:<3}", end=' ')
    print(f"合计={total}")
    print(f"    备注: {note}")